import os.path

import openpyxl
from openpyxl import load_workbook, Workbook

wb = load_workbook("Customer.xlsx")
pointer = wb.active
l = pointer["A"]
wb = Workbook()
for i in l:
    wb.save(os.path.dirname(__file__)+"/%s.xlsx" % i.value)
