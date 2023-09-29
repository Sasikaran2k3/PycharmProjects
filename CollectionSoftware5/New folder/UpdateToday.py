import datetime
import openpyxl
from openpyxl import load_workbook, Workbook

date = "".join(str(datetime.date.today()).split("-"))

try:
    wb = load_workbook(date + ".xlsx")
    pointer = wb.active
except FileNotFoundError:
    print(date)

list_of_shop = pointer["C"]
list_of_amt = pointer["D"]

for i in range(len(list_of_shop)):
    try:
        add_wb = load_workbook(list_of_shop[i].value + ".xlsx")
        add_pointer = add_wb.active
        dates = add_pointer["A"]
        amts = add_pointer["B"]
        max_val = len(amts)
        for j in range(1,len(dates)+1):
            cell = add_pointer.cell(row=j,column=1)
            if cell.value == None:
                add_pointer.cell(row = j, column=1).value = str(datetime.date.today())
                add_pointer.cell(row=j, column=2).value = list_of_amt[i].value
                break
            elif str(datetime.date.today()) in cell.value:
                add_pointer.cell(row=j,column=2).value = list_of_amt[i].value
                break

        else:
            add_pointer.cell(row=j+1, column=1).value = str(datetime.date.today())
            add_pointer.cell(row=j+1, column=2).value = list_of_amt[i].value
        """
        if dates[0].value == None:
            print("if")
            add_pointer["A1"].value =str(datetime.date.today())
            add_pointer["B1"].value = list_of_amt[i].value
        elif str(datetime.date.today()) in dates[-1].value:
            print("elif")
            add_pointer["B%d"%max_val].value = list_of_amt[i].value
        else:
            print("else")
            add_pointer["A%d" % max_val+1].value = str(datetime.date.today())
            add_pointer["B%d" % max_val+1].value = list_of_amt[i].value"""
        #add_pointer.append((str(datetime.date.today()), list_of_amt[i].value))
        add_wb.save(list_of_shop[i].value + ".xlsx")
    except Exception as e:
        print(list_of_shop[i].value)
        print(e)