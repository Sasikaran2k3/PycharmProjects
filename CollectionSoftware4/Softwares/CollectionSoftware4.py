import datetime
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
import csv

import openpyxl
from openpyxl import load_workbook

root = Tk()
root.title("Ravi Oil Store")
img = PhotoImage(file='ros_logo.png')
root.iconphoto(False, img)
root.geometry("1080x720")
area_list = ["Custom","MMDA","Choolaimedu", "Anna Nagar", "Full", "Arumbakkam", "Kilpak", "Local"]
Choolaimedu = open("CHOOLAIMEDU.txt", "r").read().split("\n")
anna_nagar = open("ANNA_NAGAR.txt","r").read().split("\n")
mmda = open("MMDA.txt","r").read().split("\n")
arumbakkam = open("ARUMBAKAM.txt","r").read().split("\n")
custom = open("CUSTOM.txt","r").read().split("\n")
kilpak = open("KILPAK.txt","r").read().split("\n")
local = open("LOCAL.txt","r").read().split("\n")
l = ["ARI ALAGAR STORES", "ARIHARAN STORES", "ARTHI FISH STORE", "ARTHI STORE(ARUM)", "ARUL HOTEL", "ARUL STORE (CHOO)", "ARUL SUPER STORE", "ARUN STORES", "ARUN STORES(KOYM)", "ARUNACHALA TRADERS", "ASHVINI PROVISION", "ASHWIN PALAMUDIR NILAYAM", "ASR VELUS PAZHAMUDIR CHOLAI & SM (NOLAM)", "ASR VELUS PAZHAMUDIR CHOLAI & SUPER MARKET", "AST PAZHAMUDHIR NILAYAM", "ATHITHI STORE", "AVG REDDY", "AVP CHANDRA STORE", "AVS STORE", "AYYANAR OIL STORE", "AYYANAR STOER (K-87)", "AYYANAR STORE (G-87)", "AYYANAR STORE (GA307)", "AYYANAR STORE (KIL)", "AYYANAR STORE (THIRU)", "AYYANAR STORE (Y-20)", "AYYANAR STORE NSK", "AYYANAR STORE(CHOOLAI)", "AYYANAR STORE(NAVEL)", "AYYANAR TRADERS(GB-74)", "AYYAPPA STORE (ANNA NAGAR)", "AYYAPPA STORE(SHENOY)", "AYYNAR STORE (NSK 1)", "BABU STORE (ARUM)", "BAGAVATHI AMMAN FRUITS", "BAKYAM STORE (CMBT)", "BAKYAM TRADERS(B-79)", "BALAJI OIL STORE", "BALAJI STORE(AYAN)", "BALAJI STORE(KOYM)", "SRI BALAMURUGAN STORE", "no use BALAMURUGAN STORE(ANNA)", "BARANI STORE", "BARANI STORE(59)", "BHARANI BHAVAN", "BHARANI FAST FOODS", "BISMILLAH STORE", "BJ ENTERPRISES", "BKR STORE", "BR AGENCY", "CHANDRA STORE", "CHANDRA STORE MMDA", "CHANDRA STORE(KOYM)", "CHANDRA SUPER MARKET", "CHENNAI FOODS", "CHENNAI HERITAGE HOSPITAL", "CHENNAI SUPER STORE", "CHENTHUR MURUGAN TRADERS", "CHIDAMBARAM STORE", "CHOICE PROVISION STORE", "COESS VENTURES", "D.S.NAIDU STORE", "DAILY DELITE SUPER MARKET", "DAILY NEEDS", "DEEPAM FOODS(THIRU)", "DEIVAM EVERGREEN SUPERMARKET PVT LTD", "DEVAR PAZHAMUDHIR NILAYAM", "DEVI KARUMARI STORE", "DHANALAKSHMI STORE", "DHANASEI", "DINDUKAL BIRYANI MESS", "DIVIYA STORE", "DIVYA TRADERS", "DURGA (NATTU)", "ESSENTIAL4U", "ESSENTIAL4U(ANNA)", "ESWARI OIL STORES", "EVERSHINE ENTERPRISE", "EVERY DAILY HOME NEEDS", "FRESH 2 DAY", "FRESH OF FINE GERERAL STORE", "FRESH SUPERMARKET", "GANAPATHI BHAVAN", "GANDHIRAJAN STORE (GSK)", "GANESH BHAVAN HOTEL(VP)", "GANESH STORE (NEW AVD RD)", "GANESH STORE(13TH)", "GANESH STORE(LOTUS)", "GANESH STORE(N.S.K.)", "GANESH STORE(SHENOY)", "GANGA SUPER STORE", "GANGAS TIFFEN CENTRE", "GANI STORE", "GIRIJAPAATI FOODS PVT LTD", "GNANAM STORE", "GOOD LIFE ENTERPRISES", "GOOD LIFE SUPER MARKET", "GOOD LUCK STORE", "GOODLIFE SUPERMART (ANNA)", "GR PROVISION", "GREEEN LAND BAKERY", "GREEN MARKETING", "GROCERMAIL E-COMMERCE VENTURES PRIVATE LIMITED", "GSK SUPER MARKET", "HANEEFA SUPER MARKET", "HARI KRISHNA STORE", "HARI OM PHARMA", "HARICHANDRA TRADERS", "HARIYANTH PROVISION", "HOME NEEDS", "HOT CHIPS & SNACKS", "HOT CHIPS & SNACKS (NEW)", "HOT CHIPS (KIL)", "HOT CHIPS(ANNA)", "HOT CHIPS(CHOO)", "HOTEL GANAPATHY BHAVAN", "HOTEL SRI RAJ BHAVAN", "IBRAHIM STORE", "INDIRA STORE (11 TH)", "INDIRA STORE(8TH)", "ISHWARYA SWEETS", "ISMAIL STORE (NELSON)", "ISMAIL STORE(KIL)", "JAFFER STORE", "JAI AKASH STORE", "JAI DURGA STORE", "JAMAILA STORE", "JAMAL MOIDEEN STORE", "JASMINE SUPER MARKET", "JAYA LAKSHMI SWEETS", "JAYABAL STORE", "JAYADURGA STORE", "JAYALAKSHMI STORE(ARUM)", "JAYALAKSHMI SUPER  MARKET", "JAYAMANI STORE", "JAYARAM STORE", "JJ STORE(BAAI)", "JOSEPH STORE", "JUPITER SUPER MARKET", "KALA HOT CHIPS", "KALA STORE", "KALA SWEETS BAKEREY", "KAMATCHI STORE (ARUM)", "KAMATCHI STORE (KOYAM)", "KAMATCHI STORE (MMDA)", "KAMATCHI STORE(ANNA)", "KANAKADHAARA FRUITS AND VEGETABLES", "KANCHI FOODS", "KANDAN OIL STORE", "KANDAN VEG MART", "KANNAN STORE (ANNA)", "KANNIYAPPAN VEGETABLE", "KARIYA KALLIAMMAN FRUITS", "KARPAGA VINAYAKAR FRESH FRUIT", "KARPAGAM FOODS", "KARPAGAMBAAL FRESH FRUITS (ADYAR)", "KARPAGAMBAAL FRESH FRUITS (VALLUVARKOTTAM)", "KAVINS STORE", "KAVITHA HOSTEL", "KAVITHA STORE", "KAVITHA STORE (CHOO)", "KAVIYA & KAVITHA STORE", "KERALA HOT CHIPS & SNACKS", "KERALA HOT CHIPS & SNACKS (TAMBARAM)", "KERALA HOT CHIPS (SANTHI)", "KERALA HOT CHIPS AND SNACKS(COIM)", "KILPAUK GARDEN VEGEFURITS MARKET", "KING'S SUPER MARKET", "KIRTHIKA STORE", "KITCHA STORE (GB-84)", "KITCHAA STORE", "KOVAI PAZHAMUDHIR NILAYAM  (AMMAN)", "KPN FARM FRESH PRIVATE LIMITED(VANAGARAM)", "KRISHNA CAFE", "KSM CHETTIYAR CHEKKU OIL MILL", "KUMARAN STORE(CHOO)", "KUMARAN STORE(METHA)", "KURINJI TRADING POINT", "LAKSHMAN & LAKSHMI SWEETS", "LAKSHMI & CO", "LAKSHMI & CO (NEELANKARAI)", "LAKSHMI STORE (KOYM)", "LAKSHMI STORE (RV NAGAR)", "LAKSHMI STORE(VOC)", "LAKSHMI SWEETS", "LAKSHMI SWEETS (CHOO)", "LIMRAS ENTERPRISES (P) LTD", "LINGAM STORE", "LINGAM STORE (LOCAL)", "LINGAM STORE(MMDA)", "LUCKY STORE", "M. BAKKYAM STORE", "MAAHE SHOPPEE", "MADAN TRADERS", "MADINA TRADE POINT", "MADRAS -10", "MADRAS AGRO PRODUCTS", "MAHALAKSHI STORE", "MAHALAKSHMI ENTERPRISES", "MAHALAKSHMI OIL TRADERS", "MAHALAKSHMI PROVISION", "MAHALAKSHMI STORE", "MAHALAKSHMI STORE(CHOO)", "MAHARAJA STORE(CHOO)", "MANI TRADERS", "MANIGANDAN STORE", "MANO TRADERS", "MARGIN FREE SUPER MARKET", "MARINAKART", "MATHAJI STORE", "MEENAKSHI TRADERS", "MEGA MART", "MERCY HOME", "METRO GRAND HOTEL", "METRO SUPER MARKET", "MINAR BIRIYANI HOTEL", "MIRCHI HOME", "MIRORA HOTELS AND RESTAURANRTS (P).LTD", "MITIS SNACKS", "MOOGAMBIGAI STORE", "MURALI STORES", "MURUGAN STORE (18TH)", "MURUGAN STORE (AK 1)", "MURUGAN STORE (ARUM)", "MURUGAN STORE (G-87)", "MURUGAN STORE (K4)", "MURUGAN STORE (TP)", "MURUGAN STORE K4", "MURUGAN STORE(AYYA)", "MURUGAN STORE(JAI)", "MURUGAN STORE(KANDAN)", "MURUGAN STORE(KOYM)", "MURUGAN STORE(SHENOY)", "MURUGAN STORES SUPER MARKET (RAJAN)", "MURUGAN SWEETS", "MUTHU STORE", "MUTHUKRISHNA STORE", "MUTHUMALAI AMMAN STORE", "MUTTAI KADAI", "MY SHOP", "N.SELVAM (KMC)", "N.SUGANTHI (KMC)", "NALAS APPA KADAI", "NANDAN STORE", "NANDHINI TRADERS", "NATARAJA STORE", "NATIONAL STORE", "NATURE BOX TODAY", "NAVASAKTHI VINAYAGAR FRUITS", "NEW AYYANAR STORE", "NEW CHIDAMBARAM STORE", "NEW MAHE STORE(0)", "NEW NATIONAL STORE", "NEW PANDIAN STORE", "NEW PRABU STORE (GA32)", "NEW RAJA STORE", "NEW SAKTHI STORE", "NEW SHANTHI STORE", "NEW SHREE AYYANAR OIL STORE", "NEW THIRUMANI STORE", "NITHYA AMIRTHAM INDIAN FOOD PRIVATE LIMITED", "NITHYAM FOODS", "NSK HOT CHIPS", "NUTS 'N' SPICES", "OM MURUGA STORE", "OM STORE", "OM STORES(KOYM)", "P.R. AGENCIES", "P.S.P.FOODS AND GRAINS", "PADMA STORE", "PADMANABAN STORE", "PADMAVATHI SWEETS", "PALLAVI STORE", "PALRAJAN STORE", "PANDIAN STORE (ARUM)", "no use PANDIAN STORE (F-53)", "PANDIAN STORE (H-87)", "PANDIAN TRADERS", "PANDIAN TRADERS (F-53)", "PANDIAN TRADERS (GC-25)", "PANDIYAN TRADERS (D-86)", "PANJALI STORE", "PATHRAKALI STORE", "PATHUBANATHAN STORE", "PATTU RAJA STORE", "PATTU STORE (GA-36)", "PATTU STORE (GA-36)", "PAZHAMUDHIR NILAYAM", "PAZHAMUDIR NILAYAM FARM FRESH (GOPALAPURAM)", "PAZHAMUDIR NILAYAM FARM FRESH(MANDAVELI)", "PERIYANDAVAR OIL STORE", "PERIYANDAVAR STORE", "PERIYAPERATTI STORE", "PIECE OF CAKE", "PKR STORE", "PONMANI SWEETS", "PONNAIAMMAN STORE", "PONNUSAMY HOTEL", "PRABHA STORE MMDA", "PRABU STORE (A-121)", "PRABU STORE (GA-32)", "PRAGATI FOODS", "PRAKASH STORE", "QUALITY HOME", "QUIET SUPER MARKET", "R.K. NITISH STORE", "R.V TRADERS", "RABBANA STORE", "RAGAVENDRA STORE", "RAJA OIL STORE", "RAJA STORE (ARUM)", "RAJA STORE(ANNA)", "RAJA STORE(KOYM)", "RAJAM STORE", "RAJAMMAL ENTERPRISES", "RAJAPPAN STORE", "RATHI STORE", "RBN STORES", "REDDY ANDHRA MESS", "RENGA AGENCY", "RENGA STORE", "REVATHI STORE", "REVATHI SUPER MARKET", "ROBERT STORE(TP)", "RR AGENCY", "RT GURUSAMY STORE", "RV STORE", "S.S. GREEN BASKET", "S.S.S.STORE", "SA SWAMY STORE", "SADHANA TRADERS", "SAI PROVISION STORE", "SAI RAM CHIPS", "SAI RAM HOT CHIPS", "SAI SUPER MARKET", "SAIS PROVISION STORE", "SAKTHI STORE (ANNA )", "SAKTHI STORE (KAT)", "SAKTHI SUPER MARKET", "SAKTHI VINAYAGAM STORE", "SAMBATH STORE", "SANABEL BRIYANI POINT", "SANDHYA BHAVAN HOTEL", "SANKAR STORE", "SANKARAN AGENCIES", "SANTHOSH STORE(ARUM)", "SANTHOSH SUPER MARKET(ANNA)", "SANTHOSH SUPER STORE", "SANTHOSH SUPER STORE(18TH)", "SARAN SWEETS", "SARANYA STORE(CHOO)", "SARASWATHI RICE STORE", "SARAVANA AGENCIES", "SARAVANA STORE (18 TH)", "SARAVANA STORE (KOYM)", "SARAVANA STORE (PARK RD)", "SARAVANA STORE (Y 20 )", "SARAVANA STORE(18TH)", "SARAVANA STORE(KANDAN)", "SARAVANA STORE(TP)", "SARAVANAS SUPER MARKET", "SASIKUMAR STORE", "SATHANA TRADERS", "SCARLATE SUPER MARKET", "SEDHU STORE", "SEKAR STORE", "SEKAR STORES (AMMAN)", "SELLAPPA HIGH QUALITY VEG", "SELLIAMMAN STORE", "SELVA GANESH STORE", "SELVAI STORE", "SELVAKUMAR PROVISION STORE", "SELVAKUMAR SUPER MARKET", "SELVAM OIL STORE (NEW)", "SELVAM STORE", "SELVAM STORE (7TH)", "SELVAM STORE (MADURAVOYAL)", "SELVAM STORE (NEDUVAKARAI)", "SELVAM STORE (NSK)", "SELVAM STORE(ANNA)", "SELVAM STORE(ARUM)", "SELVAM STORE(LAKSHMAN)", "SELVAM STORE(THANGAM)", "SELVAM STORE(THIRU MULAR)", "SELVAM SWEETS & BAKERY(SHENOY)", "SELVAM SWEETS BAKERY AND SNACKS", "SELVI SUPER MARKET", "SELVI SUPER MARKET (SHENOY)", "SENTHIL ANDHAVAR FRUITS", "SENTHIL ANDHAVAR STORE", "SENTHIL MURUGAN STORE", "SENTHIL TRADERS", "SENTHIL VEL STORE", "SH STORE", "SHAHI FOODS", "SHANDHI OIL MILLS", "SHANMUGASUNDARAM STORE", "SHANTHI OIL", "SHOPPING SINGAPPORE (NEW)", "SHREE AKHILANDESWARI AMMAN FRUITS", "SHREE RAJAVELL TRADERS", "SHRI SABARI AGENCY", "SHRI SARASWATHI AMMA FARM FRUITS", "SHRIYAA CHATS AND FOODS", "SIDDHA LADIES HOSTEL", "SIDDHESWARA & CO", "SIDDHI KRIPA", "SIVA RAMA TRADERS", "SIVA RATHINAM TRADERS", "SIVA STORE (10TH)", "SIVA STORE (NSK)", "SIVA STORE (THANGA COLONY)", "SIVAN STORE", "SIVARATHINAM STORE(GA-34)", "SIVAYOGAM & CO", "SK STORE", "SKS MINI SUPER MARKET", "SM DURAI RAJ STORE(A-86)", "SM DURAI RAJ STORE(E87)", "SM DURAIRAJ STORE(A-91)", "SM TRADERS", "SNEHA STORE", "SNV VENTURES LLP", "SORNA LAKSHMI", "SOWBAGHYA", "SOWTHA BAKERY", "SOWTHA STORE", "SR DEPARTMENT STORE", "SREE BHARANI FAST FOOD", "SREE JAYALAKSHMI STORE", "SREE KALA HOT CHIPS", "SREE KARAN ENTERPRISES", "SREE KRISHNA PROVISION STORE", "SREE LATHA STORE", "SREE MAHALAKSHMI MARKETING", "SREE NIVAS UDUPI HOTEL", "SREE RADHA & CO", "SREE RAJARAJESWARI RICE TRADING COMPANY", "SREE RAM TRADERS", "SRI AKSHAYAAS HOTEL", "SRI AYYANAR STORE (D27)", "SRI AYYANAR TRADERS (37)", "SRI AYYANAR TRADERS(GA307)", "SRI BALAJI STORE", "SRI BALAJI TRADING COMPANY", "SRI BALASUBRAMANIYAR TRADERS", "SRI BANNARI AMMAN FRESH FRUIT", "SRI CATERING", "SRI GANESH STORE (JAI)", "SRI HARI HARAN ENTERPRISES", "SRI JAI DURGA STORE", "SRI KANDAN OIL STORE", "SRI KRISHNA PROVISION", "SRI LAKSHMI STORE", "SRI MAHALAKSHMI STORE", "SRI MURUGAN ENTERPRISES", "SRI MURUGAN STORE (CHOO)", "SRI MURUGAN STORE(SHENOY)", "SRI PAZHANI STORE", "SRI RAGAVENDRA STORE", "SRI RAGAVENDRA TRADERS", "SRI RAMAJAYAM STORE", "SRI RANGA TRADERS", "SRI SAI SABARI STORE", "SRI SAKTHI STORE", "SRI SARAPAR MURUGAN STORE", "SRI SATHYA SUPER MARKET", "SRI SRINIVASA STORE", "SRI VARI HOT CHIPS", "SRI VIJAYLAKSHMI STORE", "SRINITHI HOTEL", "SRINIVASA STORE", "SRINIVASAN", "SRIVARI SWEETS", "SSK TRADERS", "SUBA STORE(TP)", "SUHAIL ENTERPRISES (P). LTD.", "SUMATHI STORE(TP)", "SUN OIL TRADERS", "SUN SUPER STORE", "SUNDARAM STORE(AMINI)", "SUNDARAM STORE(ARUM)", "SUNDARI STORES", "SUPRABHA HOTEL (0)", "SYED STORE", "TAMILNADU (NATTUMARUNTU KADAI)", "TAMILNADU STORE", "TAYLORS  ROAD (VEG & FRUITS)", "THAI STORE(CHOO)", "THANGAM PROVISION STORE", "THANGAM STORE", "THANGAM STORE(CMBT)", "THANGAM STORE(KOYM)", "THANGAMANI STORE", "THANGAPANDIYAN STORE", "THANGAPPAN STORE", "THANJAI PROVISION STORE", "THE CREAMS", "THE FOODLE STORE", "THE GRAND SWEETS AND SNACKS", "THILAGAM STORE", "THIRUKUMARAN STORE", "THIRUMALA STORE", "THIRUMALA SUPER MARKET", "THIRUPATHI STORE", "THIRUVENI HOTEL", "THIRUVENI STORE", "UDUPI GANESH BHAVAN", "UDIPI SANKAR BHAVAN", "V S STORE", "V.R. CATERING SERVICE", "V.S. TRADERRS", "VAGAP STORE", "VAISHNAVI SWEETS", "VANAJA STORE", "VARADHARAJAN STORE(C-52)", "VARADHARAJAN TRADERS(GC-9)", "VARATHARAJAN STORE(E-86)", "VARATHARAJAN STORE(E-87)", "VASANT CAFE", "VASANTHA BHAVAN HOTEL", "VASEER STORE", "VEERA VEG MART", "VEL STORE (ANNA)", "VELUS PAZHAMUDIR SOLAI & SUPER MARKET", "VENKATESWARA CHIPS(ANNA)", "VENKATESWARA SWEETS", "VENNILA STORE", "VETTA SUPER STORE", "VETTRII & CO", "VIGNESH CATERING SERVICE", "VIGRAMA HOTEL", "VIJAYA STORE", "VIJAYALAKSHMI STORE", "VIJAYALAKSHMI SWEETS", "VIJAYALAKSHMI TRADERS", "VIKPRI HOME NEEDS", "VINAYAGA GENERAL STORES", "VINAYAGA HOT CHIPS", "VINAYAGA HOTEL", "VINAYAGA HOTEL BHAVAN", "VISHNU PRIYA STORE", "VOICE NATURAL AND FOOD PRODUCTS", "VOPEC PHARAMACEUTICALS", "VP SHOPPING", "VPD ENTERPRISES", "VRP TRADERS", "WINWAY FOOD PRODUTS", "XAVIER STORE", "YASHI PROVISIONAL TRADERS", "YESESI SUPER MARKET", "YOVAN STORE", "ZAKIR STORES", "AK SUPER MARKET", "AV STORES", "AV STORE(2)", "A V STORES 2", "AP SELVAM STORE", "ARCHANA STORE", "ANNAI STORE(BELL)", "AROKYA BHAVAN HOTEL", "SUDALAI", "AKSHAYA TRADERS", "MURUGAN STORE (AMINJI) RAMAN", "AL-KHADER PROVSION(NEW)", "RAJA DEPARTMENTAL STORE", "KRISHNA MINI MART", "SRITHAR AGENCIES", "SIDDHESWARA & CO", "NEW SELVAKUMAR STORE(NEW)", "ARFAA HOTEL", "idhyam order", "GRB ORDER", "COFFEE HOUSE AT MOUNT ROAD", "ANANDA STORE (AYYAVO COLONY)", "MURUGAN SUPER MARKET (CHOOLAI)", "KGN HOSTEL", "PANI POORI", "ANNAM,HITAM ASSOCIATES PVT LTD", "TAYLORS  ROAD (Oil & Rice)", "amen moiden store", "RAJAN STORES", "Orange super market", "REKHAA ORGANICS (FRESH2DAY)", "FRESH N FAIR", "MANGALAM SUPER MARKET",]
area_dict = {"Choolaimedu":Choolaimedu,"Anna Nagar":anna_nagar,"Full":l,"MMDA":mmda, "Arumbakkam":arumbakkam, "Choolaimedu":Choolaimedu, "Local":local, "Kilpak":kilpak, "Custom":custom}
area_count = {"Choolaimedu":0,"Anna Nagar":0,"Full":0,"MMDA":0,"Arumbakkam":0,"Local":0,"Kilpak":0,"Custom":0}
completed = []
entry_list = []
# Need to replace datetime.date.today().strftime("%d-%m-%Y") to datetime.date.today().strftime("%d-%m-%Y").strftime("%d-%m-%Y")

def update_area():
    global Choolaimedu
    global anna_nagar
    global mmda
    global custom
    global arumbakkam
    global kilpak
    global local
    global area_dict
    Choolaimedu = open("CHOOLAIMEDU.txt", "r").read().split("\n")
    anna_nagar = open("ANNA_NAGAR.txt", "r").read().split("\n")
    mmda = open("MMDA.txt", "r").read().split("\n")
    arumbakkam = open("ARUMBAKAM.txt", "r").read().split("\n")
    custom = open("CUSTOM.txt", "r").read().split("\n")
    kilpak = open("KILPAK.txt", "r").read().split("\n")
    local = open("LOCAL.txt", "r").read().split("\n")
    area_dict = {"Choolaimedu": Choolaimedu, "Anna Nagar": anna_nagar, "Full": l, "MMDA": mmda,
                 "Arumbakkam": arumbakkam, "Choolaimedu": Choolaimedu, "Local": local, "Kilpak": kilpak,
                 "Custom": custom}
def start_combo(e):
    update_area()
    select_shops.delete(0, END)
    for i in area_dict[drop_down.get()]:
        if len(i) > 1 :
            select_shops.insert(END, i)

def get_list(e):
    total.config(text="Total collection of %s: %d" % (drop_down.get(), area_count[drop_down.get()]))
    l = area_dict["Full"]
    out = []
    leng = len(str(data.get()))
    if data.get() != "":
        for i in l:
            if(data.get().upper().startswith(i[:leng])):
                out.append(i)
        list_box.delete(0,END)
    else:
        out = l[:]
        list_box.delete(0, END)
    for i in out:
        list_box.insert(END,i)
        if i in completed:
            list_box.itemconfig(END,bg ="green",selectbackground="green")
    global store_name
    store_name=list_box.select_set(0)

def select_shop(ev):
    select_shops.insert(END, list_box.get(list_box.curselection()[0]))
    text.delete(0,END)
    escape_operation(ev)

def enter(e):
    list_box.focus_set()


def change_value(ev):
    output_frame.grid_slaves(row=1,column=3)[0].focus_set()

def total_calculator(ev):
    def find_selects():
        cash = 0
        on = 0
        for i in entry_list:
            if 'selected' in output_frame.grid_slaves(row = i, column=6)[0].state():
                amt = output_frame.grid_slaves(row = i,column= 3)[0].get()
                online_pay = output_frame.grid_slaves(row = i,column= 4)[0].get()
                if amt.strip().isnumeric():
                    cash += int(amt)
                if online_pay.strip().isnumeric():
                    on += int(online_pay)
        area_count[drop_down.get()] = cash + on
        total.config(text="Cash : %d, Online : %d, Total : %d" % (cash, on, cash+on))
    print("\n-----------------")
    find_selects()
    print("----------------------\n")

def show_accounts(ev):
    short = list(set(select_shops.get(0,END)))
    short.sort()
    if short == []:
        short = area_dict[drop_down.get()]
    else:
        f = open(drop_down.get() + ".txt", "w")
        for i in short:
            if i != "":
                f.write(i + "\n")
        f.close()
        print(open(drop_down.get() + ".txt", "r").read()+'written')
    select_shops.pack_forget()
    list_box.pack_forget()
    save_list.pack_forget()
    outside.pack(fill=BOTH,expand=YES)
    buttons_of_operation.pack(side=BOTTOM)
    global store_name
    global entry_list
    entry_list = []
    global counter
    counter = 0
    for i in output_frame.winfo_children():
        i.destroy()
    Label(output_frame, text="Store Name:", background='light grey', font=("Calibri", 15)).grid(row=0, column=0)
    Label(output_frame, text="Date:", background='light grey', font=("Calibri", 15)).grid(row=0, column=1, padx=70)
    Label(output_frame, text="Amount:", background='light grey', font=("Calibri", 15)).grid(row=0, column=2)
    Label(output_frame, text="Amt Received:", background='light grey', font=("Calibri", 15)).grid(row=0, column=3)
    Label(output_frame, text="Online Pay:", background='light grey', font=("Calibri", 15)).grid(row=0, column=4)
    Label(output_frame, text="Balance:", background='light grey', font=("Calibri", 15)).grid(row=0, column=5)
    for i in short:
        store_name = i
        global wb
        try:
            global wb
            wb = open(store_name+".csv","r")
        except Exception as e:
            print(e)
            continue
        reader = csv.reader(wb)
        #for i in output_frame.winfo_children():
        #    i.destroy()
        Label(output_frame, text=store_name, font=("Calibri", 15)).grid(row=counter + 1, column=0)
        date=[]
        amt=[]
        for i in reader:
            if i == []:
                continue
            date.append(i[0])
            amt.append(i[1])
        if (date,amt) == ([],[]):
            counter += 1
            continue
        for row,i in enumerate(amt):
            if date[row] == None and row != 0:
                print("empty")
                break
            if row!=0:
                Label(output_frame, text=store_name, font=("Calibri", 15)).grid(row=counter + 1, column=0)
            if i.isnumeric():
                Label(output_frame, text = date[row],font=("Calibri", 15)).grid(row=counter+1, column=1)
                Label(output_frame, text = i, font=("Calibri", 15)).grid(row=counter+1,column=2)
                e = Entry(output_frame,relief=GROOVE,font=("Calibri", 15),width=10)
                e.grid(row=counter+1, column=3, padx=35, pady=2)
                e.bind("<KeyRelease>", calculate_and_focus)
                e.bind("<Down>", movement_down)
                e.bind("<Up>", movement_up)
                e.bind("<Right>", movement_right)
                e.bind("<Return>", movement_down)
                e.bind("<Escape>",enter)

                o = Entry(output_frame, relief=GROOVE, font=("Calibri", 15),width=10)
                o.grid(row=counter + 1, column=4, pady=2, padx=35)
                o.bind("<KeyRelease>", calculate_and_focus)
                o.bind("<Down>", movement_down)
                o.bind("<Up>", movement_up)
                o.bind("<Left>", movement_left)
                o.bind("<Return>", movement_down)
                o.bind("<Escape>", enter)

                c = ttk.Checkbutton(output_frame, state='selected')
                # c.bind("<ButtonRelease-1>",total_calculator)
                c.grid(row=counter + 1, column=6)
                c.state(['!alternate', '!disabled', 'selected'])
                Label(output_frame, font=("Calibri", 15), text=i).grid(row=counter + 1, column=5)
                counter += 1
                entry_list.append(counter)
            else:
                counter += 1
    e.bind("<Return>",jump_to_save)
    output_frame.pack(fill = BOTH,expand=YES,padx=2,pady=2)
    canva.create_window((canva.winfo_width()//2, 0), window=output_frame, anchor=CENTER)
    canva.update_idletasks()
    canva.configure(scrollregion=canva.bbox(ALL))
    canva.yview_moveto(0)
    #wait_box(1)
    e.bind("<Return>", jump_to_save)

def recreate(e):
    outside.pack_forget()
    buttons_of_operation.pack_forget()
    list_box.pack(fill=X)
    select_shops.pack()
    save_list.pack()
def calculate_and_focus(ev):
    global counter
    canva.yview_moveto(float((output_frame.focus_get().grid_info().get('row') - 2) / counter))
    new = output_frame.focus_get().grid_info()['row']
    val = output_frame.grid_slaves(row=new,column=3)[0].get()
    on = output_frame.grid_slaves(row=new,column=4)[0].get()
    val = int(val) if val.isnumeric() else 0
    on = int(on) if on.isnumeric() else 0
    print(val , on)
    bal = int(output_frame.grid_slaves(row=new, column=2)[0].cget("text")) - int(val) - int(on)
    if (val,on) == (0,0) :
        output_frame.grid_slaves(row=new, column=5)[0].configure(text=output_frame.grid_slaves(row=new, column=2)[0].cget("text"))
    else:
        output_frame.grid_slaves(row=new, column=5)[0].configure(text=bal, background="lightgreen")
    """
    if val.isnumeric():
        if on.isnumeric():
            
        else:
            bal = bal = int(output_frame.grid_slaves(row=new, column=2)[0].cget("text"))- int(val)
        output_frame.grid_slaves(row=new,column=5)[0].configure(text=bal,background="lightgreen")
    else:
        output_frame.grid_slaves(row=new, column=5)[0].configure(text=output_frame.grid_slaves(row=new, column=2)[0].cget("text"))
"""

def movement_down(e):
    total_row =output_frame.grid_size()[1]
    info = output_frame.focus_get().grid_info()
    current_row = entry_list[entry_list.index(info["row"])+1]
    if total_row-1 >= current_row:
        output_frame.grid_slaves(current_row,column=3)[0].focus_set()


def movement_up(e):
    info = output_frame.focus_get().grid_info()
    current_row = entry_list[entry_list.index(info["row"])-1]
    if 1 <= current_row:
        output_frame.grid_slaves(current_row, column=3)[0].focus_set()

def movement_right(e):
    info = output_frame.focus_get().grid_info()
    current_row = entry_list[entry_list.index(info["row"])]
    if 1 <= current_row:
        output_frame.grid_slaves(current_row, column=4)[0].focus_set()

def movement_left(e):
    info = output_frame.focus_get().grid_info()
    current_row = entry_list[entry_list.index(info["row"])]
    if 1 <= current_row:
        output_frame.grid_slaves(current_row, column=3)[0].focus_set()

def jump_to_save(e):
    buttons_of_operation.grid_slaves(row=0, column=2)[0].focus_set()


def pop_ask(e):
    def yes_even(ev):
        print("saved")
        save_to_main()
        pop.destroy()

    def no_event(ev):
        print("not saved")
        pop.destroy()
        text.focus_set()

    def highlight_yes(ev):
        yes.focus_set()

    def highlight_no(ev):
        no.focus_set()

    global pop
    pop = Toplevel(root)
    pop.geometry("500x200")
    pop.title("Save?")
    Label(pop, text="Do you really want to save?").grid(row=0,column=1)
    yes = Button(pop,text="Yes")
    yes.bind("<Return>", yes_even)
    yes.bind("<Button-1>", yes_even)
    yes.bind("<Right>", highlight_no)
    yes.focus_set()
    yes.grid(row=1,column=1,pady=10)
    no = Button(pop,text="No")
    no.grid(row=1,column=2,pady=10)
    no.bind("<Return>", no_event)
    no.bind("<Button-1>", no_event)
    no.bind("<Left>", highlight_yes)
    pop.grid()

def save_to_main():
    history = open("HISTORY" + ".csv","a",newline="")
    his = csv.writer(history)
    data_list = []
    completed_stores = []
    for enum, i in enumerate(entry_list):
        print(output_frame.grid_slaves(row=i, column=0)[0].cget("text"), "select")
        store_name = output_frame.grid_slaves(row=i, column=0)[0].cget("text")
        global wb
        global pointer
        if store_name not in completed_stores:
            print(store_name,"w")
            wb = open(str(store_name) + ".csv","w",newline="")
            pointer = csv.writer(wb)
        else:
            print(store_name, "a")
            wb = open(str(store_name) + ".csv", "a",newline="")
            pointer = csv.writer(wb)
        bal = int(output_frame.grid_slaves(row=i, column=5)[0].cget("text"))
        if bal != 0:
            date = str(output_frame.grid_slaves(row=i, column=1)[0].cget("text"))
            pointer.writerow([date,bal])
            #print((date, bal,output_frame.grid_slaves(row=i,column=3)[0].get()))
        receive = int(output_frame.grid_slaves(row=i,column=3)[0].get()) if output_frame.grid_slaves(row=i,column=3)[0].get().isnumeric() else 0
        online = int(output_frame.grid_slaves(row=i, column=4)[0].get()) if output_frame.grid_slaves(row=i, column=4)[0].get().isnumeric() else 0
        print(receive,online)
        if receive or online:
            if drop_down.get() != "Full":
                area_count["Full"] += int(receive)
            area_count[drop_down.get()] += int(receive)
            total.config(text="Total collection of %s: %d" % (drop_down.get(), area_count[drop_down.get()]))
            print([str(datetime.date.today().strftime("%d-%m-%Y")), store_name,int(receive),online])
            his.writerow([str(datetime.date.today().strftime("%d-%m-%Y")), store_name,int(receive),online])
        wb.close()
        completed_stores.append(store_name)

    history.close()
    text.delete(0, END)
    recreate(0)
    escape_operation()
"""
        if enum<len(entry_list)-1:
            if output_frame.grid_slaves(row=entry_list[enum+1], column=0)[0].cget("text") != store_name:
                pointer.delete_cols(1, 2)
                for enuma, small_data in enumerate(data_list):
                    print(enuma,small_data)
                    pointer["A%s"%(enuma+1)].value, pointer["B%s"%(enuma+1)].value = small_data
                    print(small_data)
                    #pointer.append(small_data)
                data_list.clear()
                wb.save(store_name+ ".csv")
                print("all saved")
                completed.append(store_name)
        else:
            pointer.delete_cols(1, 2)
            for small_data in data_list:
                pointer.delete_rows(0)
                print(small_data)
                pointer.append(small_data)
            data_list.clear()
            wb.save(store_name + ".csv")
            print("all saved else")
    # 2nd Update here
    wb = load_workbook(str(store_name) + ".csv")
    history = load_workbook("HISTORY" + ".csv")
    his = history.active
    pointer = wb.active
    total_row = output_frame.grid_size()[1]
    pointer.delete_cols(2)
    insert_pointer = 1
    for i in range(1,len(entry_list)+1):
        value = output_frame.grid_slaves(row=i,column=4)[0].cget("text")
        balance = output_frame.grid_slaves(row=i,column=3)[0].get()
        if balance.isnumeric():
            print(str(datetime.date.today().strftime("%d-%m-%Y")), store_name, balance)
            his.append((str(datetime.date.today().strftime("%d-%m-%Y")), store_name, int(balance)))
            if drop_down.get() != "Full":
                area_count["Full"] += int(balance)
            area_count[drop_down.get()] += int(balance)
            total.config(text="Total collection of %s: %d" % (drop_down.get(),area_count[drop_down.get()]))
        if value != 0:
            pointer['B%d' % insert_pointer].value = int(value)
            insert_pointer += 1
        else:
            pointer.delete_rows(i)
    wb.save((store_name) + ".csv")
    history.save("HISTORY.csv")
    completed.append(store_name)
    text.delete(0,END)
    escape_operation()
"""


def escape_operation(ev=0):
    text.focus_set()
def report_store():
    l = area_dict[drop_down.get()]
    print(l)
    book = openpyxl.Workbook()
    p = book.active
    total_sum = 0
    for i in l:
        print(i)
        if i == '':
            continue
        wb = load_workbook(i+".csv")
        pointer = wb.active
        date = pointer["A"]
        amt = pointer["B"]
        for row, j in enumerate(amt):
            if type(j.value) == int:
                #print(i+":", j.value, date[row].value,"\n")
                total_sum += j.value
                p.append((i, date[row].value, j.value))
    file_name = "_"+drop_down.get()+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    print(file_name)
    p.append(("Total",str(datetime.datetime.now()),total_sum))
    book.save(str(file_name)+".csv")
    book.save("Records/"+str(file_name) + ".csv")
def updateToday(ev):
    date = "".join(str(datetime.date.today().strftime("%d-%m-%Y")).split("-"))
    try:
        wb = load_workbook(date + ".csv")
        pointer = wb.active
    except FileNotFoundError:
        print(date)

    list_of_shop = pointer["C"]
    list_of_amt = pointer["D"]

    for i in range(len(list_of_shop)):
        if list_of_shop[i].value == "Cash":
            continue
        try:
            add_wb = load_workbook(list_of_shop[i].value + ".csv")
            add_pointer = add_wb.active
            dates = add_pointer["A"]
            amts = add_pointer["B"]
            max_val = len(amts)
            for j in range(1, len(dates) + 1):
                cell = add_pointer.cell(row=j, column=1)
                if cell.value == None:
                    add_pointer.cell(row=j, column=1).value = str(datetime.date.today().strftime("%d-%m-%Y"))
                    add_pointer.cell(row=j, column=2).value = list_of_amt[i].value
                    break
                elif str(datetime.date.today().strftime("%d-%m-%Y")) in str(cell.value):
                    add_pointer.cell(row=j, column=2).value = list_of_amt[i].value
                    break

            else:
                add_pointer.cell(row=j + 1, column=1).value = str(datetime.date.today().strftime("%d-%m-%Y"))
                add_pointer.cell(row=j + 1, column=2).value = list_of_amt[i].value
            add_wb.save(list_of_shop[i].value + ".csv")
        except Exception as e:
            print(list_of_shop[i].value)
            print(e)

counter = 0
Label(root, text="WELCOME TO COLLECTION SOFTWARE").pack(side=TOP,pady=5)

input_frame = Frame(root)
Label(input_frame,text= "Store Name : ").grid(row=0,column=0)
data = StringVar()
text = Entry(input_frame,width=90,bd=15,textvariable=data)
text.bind("<KeyRelease>",get_list)
text.bind("<Return>", enter)
text.bind("<Down>", enter)
text.grid(row=0,column=1)
text.focus_set()

drop = StringVar(root)
drop.set(area_list[0])
drop_down = Combobox(input_frame,value=area_list,font=("Calibri", 15))
drop_down.current(0)
drop_down.grid(row=0, column=2,padx=10)
drop_down.bind('<<ComboboxSelected>>', start_combo)
drop_down.bind('<Escape>', escape_operation)

addToday = Button(input_frame,text="Update Today")
addToday.bind("<Button-1>", updateToday)
addToday.grid(row = 0, column = 3, padx=10)

input_frame.pack(pady=10)


list_box = Listbox(root, height=10,font=("Calibri", 15))
list_box.pack(fill='x',pady=5)
#list_box.bind("<KeyRelease>",show_accounts)
list_box.bind("<Double-Button-1>", select_shop)
list_box.bind("<Return>", select_shop)
list_box.bind("<Escape>", escape_operation)
store_name = ""

select_shops = Listbox(root,font=("Calibri", 15), width=75)
select_shops.pack(pady=10)
select_shops.bind("<Delete>",lambda e: select_shops.delete(select_shops.curselection()[0]))

save_list = Button(text="Show Accounts")
save_list.bind("<Button-1>",show_accounts)
save_list.pack()

outside = Frame(root)
outside.pack(fill=BOTH,expand=YES)

canva = Canvas(outside)
canva.pack(side=LEFT,fill=BOTH,expand=YES)

scrollbar = Scrollbar(outside,orient=VERTICAL,command=canva.yview)
scrollbar.pack(side=RIGHT,fill=Y)

canva.config(yscrollcommand=scrollbar.set)
canva.bind("<Configure>",lambda e: canva.configure(scrollregion=canva.bbox(ALL)))

output_frame = Frame(canva,highlightbackground="black", highlightthickness=2)


buttons_of_operation = Frame(root)
report = Button(buttons_of_operation, text="Report", command=report_store).grid(row=0,column=0,padx=15)

total = Label(buttons_of_operation,text = "Total collection of %s: %d" % (drop_down.get(),area_count[drop_down.get()]),font=("Calibri", 15))
total.grid(row=0,column=1,padx=15)

total_but = Button(buttons_of_operation, text="Total", width=10)
total_but.bind("<Return>",total_calculator)
total_but.bind("<Button-1>",total_calculator)
total_but.grid(row=0, column=2)


save_but = Button(buttons_of_operation, text="Save Changes")
save_but.bind("<Return>", pop_ask)
save_but.bind("<Button-1>", pop_ask)
save_but.grid(row=0,column=3,padx=25)

back_but = Button(buttons_of_operation,text="Back")
back_but.bind("<Button-1>",recreate)
back_but.grid(row=0,column=4,padx=25)

buttons_of_operation.pack(side=BOTTOM,pady=(0,50))


lab = Label(root,text="")
lab.pack()

root.mainloop()
