import sys
import time
import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By


def unReadMsg():
    unReadMsg = WebDriverWait(browser, 1000).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'div[class="_1pJ9J"]')))
    unReadMsg.click()


def lastMsg():
    time.sleep(5)
    try:
        LastMsg = browser.find_elements(By.CSS_SELECTOR, 'div[class="_2wUmf message-in focusable-list-item"]')[-1].text.split('\n')[0]
    except:
        print("LeftClick Error ... Maybe")
        browser.refresh()
    # This LastMsg returs 2 string seperated by \n so 1st line has actual data and 2nd line with timestamp
    myLastMsg = browser.find_elements(By.CSS_SELECTOR, 'div[class="_2wUmf message-out focusable-list-item"]')#[-1].text.split('\n')[0]
    if myLastMsg != []:
        myLastMsg = myLastMsg[-1].text.split('\n')[0]
    print(myLastMsg)
    return [LastMsg, myLastMsg]


def textAndReply(query):
    TBox = browser.find_element(By.CSS_SELECTOR, 'p[class="selectable-text copyable-text"]')
    if query[0]=='bye':
        sys.exit()
    if query[1] == []:
        TBox.send_keys(intro)
        return 0
    if 'S' in query[1]:
        data = SProcessor(query)
    else:
        data = QProcessor(query)
    if 'D' in query[1]:
        data = DProcessor(query)
    TBox.send_keys(data)
    #reply = stepReplyOutput(stepNo)
    #TBox.send_keys(reply)


def PressFormate(part):
    partion=''
    for i in range(0, len(part)):
        partion+="Press "+str( (i+1))+ " for "+str(part[i])+ "\n"
    return partion
def CloseChat():
    browser.find_element(By.CSS_SELECTOR,'div[data-tab="6"][title="Menu"]').click()
    browser.find_element(By.CSS_SELECTOR,'div[aria-label="Close chat"]').click()



def getDataFromSheet():
    contacts=[]
    for i in range(1,sheet.max_row+1):
        contacts.append(sheet["A"+str(i)].value)
    contactNo =browser.find_element(By.CSS_SELECTOR,'div[class="_24-Ff"]').text.split("\n")
    if contactNo[0] not in contacts:
        print("Appending")
        row=sheet.max_row+1
        sheet["A"+str(row)].value = contactNo[0]
        workspace.save("DetailsStorage.xlsx")
        return row
    else:
        pos = contacts.index(contactNo[0])+1
        return pos

def DProcessor(query):
    genDetail = "If You are not sure about the particular detail\nThen Enter NA to skip . "
    if query[1] == "D1": # ProductDetail
        if query[0].replace(" ","").isalpha():
            columnLetter = "G"
            stepExcelInclude(columnLetter, query[0])
            return "Accepted\nYou have to sumbit some document related to the case LATER\nEnter the Document Description :\nD2\n"
        else:
            return "Invalid\nKindly Re-Enter without digits and special character\nD1\n"
    if query[1] == "D2":
        columnLetter = "G"
        if len(query[0])>10 and len(query[0])<100: #Doc Description
            stepExcelInclude(columnLetter,query[0])
            return "Accepted\nEnter the Brand name:\n%s\nD3\n"%genDetail
        else:
            return "Invalid\nToo length or Too short(min-10 letters)\nD2\n"
    if query[1] == "D3":# Brand Name
        columnLetter = "I"
        if query[1].strip().upper() == "NA":
            stepExcelInclude(columnLetter,"NA")
            return "Noted NA\nD4\n"
        elif len(query[0])<100:
            stepExcelInclude(columnLetter, query[0])
            return "Accepted\nD4\n"
        else:
            return "Invalid\nToo long(max - 100)\nD3\n"

def SProcessor(query):
    if query[1] == "S":
        columnLetter = "D"
        if query[0].isnumeric():
            if int(query[0]) >= 1 and int(query[0]) <= 37:
                s = State[int(query[0])-1]
                stepExcelInclude(columnLetter, s)
                Dist = PressFormate(listOfDist[int(query[0])-1])
                return "U have selected %s\n%s\nS%s\n"%(s,Dist,query[0])
            else:
                return "Invalid\n"+States
        else:
            return "Invalid\n" + States
    if "S" in query[1] and "_" not in query[1]:
        columnLetter = "E"
        code = int((query[1]).replace("S",""))-1
        if query[0].isnumeric():
            query[0] = int(query[0])
            if int(query[0]) >= 1 and int(query[0]) <= len(listOfDist[code]):
                print("limit %d"%len(listOfDist[code]))
                dist = listOfDist[code][query[0]-1]
                stepExcelInclude(columnLetter, dist)
                code = str(code+1)+"_"+str(query[0])
                print(code)
                Sub = PressFormate(getSubDistricts(query))
                return "U have selected %s\n%s\nS%s\n"%(dist,Sub,code)
            else:
                return "Invalid\n"+PressFormate(listOfDist[code])+query[1]+"\n"
    if "S" in query[1] and "_" in query[1]:
        code = query[1].split("_")
        temp = code[0]
        code[0] =code[1]
        code[1] =temp
        leng = len(getSubDistricts(code))
        if int(query[0]) >= 1 and int(query[0]) <= leng:
            columnLetter = "F"
            subDist = getSubDist(query)
            stepExcelInclude(columnLetter, subDist)
            return "U have selected %s\nEnter the Product name :\nD1\n"%subDist
        else:
            Sub = PressFormate(getSubDistricts(code))
            return "Invalid\n"+Sub+"\n"+query[1]+"\n"
def fileGet(code):
    path = r"SubDist/D" + code + ".txt"
    f = open(path,"r")
    content = f.read()
    script = "global l\nl = " + content
    exec(script)
    return l
def getSubDistricts(query):
    code = query[1].replace("S","")
    content = fileGet(code)
    return (content[int(query[0])-1])
def getSubDist(query):
    code = query[1].replace("S","")
    code = code.split("_")
    content = fileGet(code[0])
    dist = content[int(code[1])-1]
    return dist[int(query[0])-1]

def QProcessor(query):
    if query[1] == "Q1":
        if query[0] == "1":
            return "U Have selected File Complaint .\n"+catagory
        elif query[0] == "2":
            return "U Have selected Track Complaint .\n"
        elif query[0] == 3:
            return "U Have Know ur rights .\n"
        else:
            return "Invalid.\n"+intro
    if query[1] == "Q11":
        columnLetter="B"
        if query[0] == "1":
            stepExcelInclude(columnLetter,query[0])
            return subCatOfPack
        elif query[0] == "2":
            stepExcelInclude(columnLetter, query[0])
            return subCatOfCate
        elif query[0] == "3":
            stepExcelInclude(columnLetter, query[0])
            return subCatOfOn
        elif query[0] == "4":
            stepExcelInclude(columnLetter, query[0])
            return subCatOfRet
        elif query[0] == "5":
            stepExcelInclude(columnLetter, query[0])
            return subCatOfOth
        else:
            return "Invalid\n"+catagory
    if query[1] == "Q12":
        columnLetter = "C"
        if query[0].isnumeric():
            if int(query[0]) >= 1 and int(query[0]) <= 17:
                stepExcelInclude(columnLetter, query[0])
                return "SubCategory selected\nEnter the Product Name:\nM1\n"
            else:
                return "Invalid\n"+subCatOfPack
        else:
            return "Invalid\n" + subCatOfPack
    if query[1] == "Q21":
        columnLetter = "C"
        if query[0].isnumeric():
            if int(query[0])>=1 and int(query[0])<=10:
                stepExcelInclude(columnLetter, query[0])
                return "SubCategory selected\n%s\n"%States
            else:
                return "Invalid\n"+subCatOfCate
        else:
            return "Invalid\n" + subCatOfCate
    if query[1] == "Q31":
        columnLetter = "C"
        if query[0].isnumeric():
            if int(query[0])>=1 and int(query[0])<=3:
                stepExcelInclude(columnLetter, query[0])
                return "SubCategory selected\n%s\n"%States
            else:
                return "Invalid\n"+subCatOfOn
        else:
            return "Invalid\n" + subCatOfOn
    if query[1] == "Q41":
        columnLetter = "C"
        if query[0].isnumeric():
            if int(query[0])>=1 and int(query[0])<=5:
                stepExcelInclude(columnLetter, query[0])
                return "SubCategory selected\n%s\n"%States
            else:
                return "Invalid\n"+subCatOfRet
        else:
            return "Invalid\n" + subCatOfRet
    if query[1] == "Q51":
        columnLetter = "C"
        if query[0].isnumeric():
            if int(query[0])>=1 and int(query[0])<=1:
                stepExcelInclude(columnLetter, query[0])
                return "SubCategory selected\n%s\n"%States
            else:
                return "Invalid\n"+subCatOfOth
        else:
            return "Invalid\n" + subCatOfOth
    workspace.save("DetailsStorage.xlsx")
def MProcessor(query):
    if query[1] == "M1":
        pass

def stepExcelInclude(columbLetter,query):
    cell = getDataFromSheet()
    print(cell, query, columbLetter + str(cell))
    write = sheet[columbLetter + str(cell)]
    write.value = query
    workspace.save(r"DetailsStorage.xlsx")
# ------------------------------------------------------------------- #

l=[]
workspace = openpyxl.load_workbook(r"DetailsStorage.xlsx")
sheet = workspace.active

opt=Options()
#options.add_experimental_option("debuggerAddress","localhost:1234")
opt.add_argument(r'--user-data-dir=E:\Hackathon\BrowserChromes\Whatsapp')
#opt.add_argument("--headless")

services = Service(r"C:\Users\HP\PycharmProjects\WebDriver\chromedriver.exe")

browser=Chrome(service=services,options=opt)

browser.get("https://web.whatsapp.com/")
print("Opened")
browser.implicitly_wait(15)

intro="Hi\nThis is a place where we help You to file a case in fssai website\nPress 1 to File a case\nPress 2 to track the case\nPress 3 to know your rights\nQ1\n"
catagory = 'Which Category Your case belongs to :\nPress 1 for Package Food\nPress 2 for Food Catering Premises\nPress 3 for Online aggregator/e-commerce\nPress 4 for Retailer Premises\nPress 5 for Others\nQ11\n'
subCatOfPack = 'Which Sub-Category Your case belongs to :\nPress 1 for Dairy products\nPress 2 for Fats & oils\nPress 3 for Edible ices including sorbet\nPress 4 for Confectionery\nPress 5 for Cereal & cereal products\nPress 6 for Bakery products\nPress 7 for Meat & meat products including poultry\nPress 8 for Fish & fish products\nPress 9 for Egg & egg products\nPress 10 for Sweetners including honey\nPress 11 for Salt, spices, soups, sauces, salads & protein products\nPress 12 for Beverages excluding dairy products\nPress 13 for Ready to eat savouries\nPress 14 for Prepared food\nPress 15 for Others\nPress 16 for Nutraceuticals\nPress 17 for Fruit and Vegetables\nQ12\n'
subCatOfCate = 'Which Sub-Category Your case belongs to :\nPress 1 for Restaurants \nPress 2 for Canteen \nPress 3 for Cafeteria \nPress 4 for Dhabas \nPress 5 for Cafe \nPress 6 for Hostel Mess \nPress 7 for Food Trucks \nPress 8 for Take aways \nPress 9 for Hotel \nPress 10 for Others \nQ21\n'
subCatOfOn = 'Which Sub-Category Your case belongs to :\nPress 1 for Prepared Food Delivering Agency \nPress 2 for Grocery Delivering Agency \nPress 3 for Others \nQ31\n'
subCatOfRet = 'Which Sub-Category Your case belongs to :\nPress 1 for Retail shops \nPress 2 for Milk & milk products retail shop \nPress 3 for Meat & meat products (including poultry & fish) retail shop \nPress 4 for Fruits & vegetable retail shop \nPress 5 for Others \nQ41\n'
subCatOfOth = 'Which Sub-Category Your case belongs to :\nPress 1 for Others \nQ51\n'
State=['Andaman And Nicobar Islands', 'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chandigarh', 'Chhattisgarh', 'Dadra & Nagar Haveli', 'Daman & Diu', 'Delhi', 'Goa', 'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jammu & Kashmir', 'Jharkhand', 'Karnataka', 'Kerala', 'Ladakh', 'Lakshadweep', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya', 'Mizoram', 'Nagaland', 'Orissa', 'Puducherry', 'Punjab', 'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana', 'Tripura', 'Uttarakhand', 'Uttar Pradesh', 'West Bengal']
States = 'Which State you belong to :\nPress 1 for Andaman And Nicobar Islands \nPress 2 for Andhra Pradesh \nPress 3 for Arunachal Pradesh \nPress 4 for Assam \nPress 5 for Bihar \nPress 6 for Chandigarh \nPress 7 for Chhattisgarh \nPress 8 for Dadra & Nagar Haveli \nPress 9 for Daman & Diu \nPress 10 for Delhi \nPress 11 for Goa \nPress 12 for Gujarat \nPress 13 for Haryana \nPress 14 for Himachal Pradesh \nPress 15 for Jammu & Kashmir \nPress 16 for Jharkhand \nPress 17 for Karnataka \nPress 18 for Kerala \nPress 19 for Ladakh \nPress 20 for Lakshadweep \nPress 21 for Madhya Pradesh \nPress 22 for Maharashtra \nPress 23 for Manipur \nPress 24 for Meghalaya \nPress 25 for Mizoram \nPress 26 for Nagaland \nPress 27 for Orissa \nPress 28 for Puducherry \nPress 29 for Punjab \nPress 30 for Rajasthan \nPress 31 for Sikkim \nPress 32 for Tamil Nadu \nPress 33 for Telangana \nPress 34 for Tripura \nPress 35 for Uttarakhand \nPress 36 for Uttar Pradesh \nPress 37 for West Bengal\nS\n'

listOfDist = [['Andamans', 'Nicobars'], ['Anantapur', 'Chittoor', 'Cuddapah', 'East Godavari', 'Guntur', 'Kadapa Municipal Corporation', 'Kakinada Mpl.Corp.', 'Krishna', 'Kurnool', 'Nellore', 'Prakasam', 'Srikakulam', 'Visakhapatnam', 'Vizianagaram', 'West Godavari'], ['Anjaw', 'Changlang', 'Dibang Valley', 'East Kameng', 'East Siang', 'Kamle', 'Kra Dadi', 'Kurung Kumey', 'Leparada', 'Lohit', 'Longding', 'Lower Dibang Valley', 'Lower Subansiri', 'Lower Subansiri', 'Namsai', 'Pakke Kessang', 'Papum Pare', 'Shi Yomi', 'Siang', 'Tawang', 'Tirap', 'Upper Siang', 'Upper Subansiri', 'West Kameng', 'West Siang'], ['Baksa', 'Barpeta', 'Biswanath', 'Bongaigaon', 'Cachar', 'Charaideo', 'Chirang', 'Darrang', 'Dhemaji', 'Dhubri', 'Dibrugarh', 'Dima Hasao', 'Goalpara', 'Golaghat', 'Hailakandi', 'Hojai', 'Jorhat', 'Kamrup Dist.', 'Kamrup Metropolitan Dist.', 'Karbi Anglong', 'Karimganj', 'Kokrajhar', 'Lakhimpur', 'Majuli', 'Marigaon', 'Nagaon', 'Nalbari', 'Sivasagar', 'Sonitpur', 'South Salmara Mancachar', 'Tinsukia', 'Udalguri', 'West Karbi Anglong'], ['Araria', 'Arwal', 'Aurangabad', 'Banka', 'Begusarai', 'Bhagalpur', 'Bhojpur', 'Buxar', 'Darbhanga', 'Gaya', 'Gopalganj', 'Jamui', 'Jehanabad', 'Kaimur (Bhabua)', 'Katihar', 'Khagaria', 'Kishanganj', 'Lakhisarai', 'Madhepura', 'Madhubani', 'Munger', 'Muzaffarpur', 'Nalanda', 'Nawada', 'Pashchim Champaran', 'Patna', 'Purba Champaran', 'Purnia', 'Rohtas', 'Saharsa', 'Samastipur', 'Saran', 'Sheikhpura', 'Sheohar', 'Sitamarhi', 'Siwan', 'Supaul', 'Vaishali'], ['Chandigarh'], ['Balod', 'Balodabazar', 'Balrampur', 'Bastar', 'Bemetara', 'Bijapur', 'Bilaspur', 'Dantewada', 'Dhamtari', 'Durg', 'Gariyaband', 'Gurela-Pendra-Marwahi', 'Janjgir - Champa', 'Jashpur', 'Kabirdham', 'Kanker', 'Kawardha', 'Kondagaon', 'Korba', 'KOREA', 'Mahasamund', 'Mungeli', 'Narayanpur', 'Raigarh', 'Raipur', 'Rajnandgaon', 'Sukma', 'Surajpur', 'Surguja'], ['Dadra & Nagar Haveli'], ['Daman', 'Diu'], ['Central', 'East', 'New Delhi', 'North', 'North East', 'North West', 'Shahdara', 'South', 'South East', 'South West', 'West'], ['North Goa', 'South Goa'], ['Ahmedabad', 'Amreli', 'Amreli Municipality', 'Anand', 'Arvalli', 'BANASKANTHA', 'Bharuch', 'Bhavnagar', 'BHUJ(KUTCHH)', 'Botad Rural', 'Chhota Udaipur', 'Dahod Municipality', 'Dang', 'Devbhumi Dwarka', 'Dohad', 'Gandhidham Municipality', 'Gandhinagar', 'Gir Somnath', 'Godhra Municipality', 'GODHRA(PANCHMAHAL)', 'HIMMATNAGAR(SABARKANTHA)', 'Jamnagar', 'Junagadh', 'Kalol Municipality', 'Mahesana', 'MAHISAGAR', 'MORBI', 'NADIAD(KHEDA)', 'Narmada', 'Navsari', 'Patan', 'Porbandar', 'Rajkot', 'Rajkot Municipal Corporation', 'Surat', 'Surat Municipal Corporation', 'Surendranagar', 'Tapi', 'Vadodara', 'Vadodara Municipal Corporation', 'Valsad', 'VYARA(TAPI)'], ['Ambala', 'Bhiwani', 'Charkhi Dadri', 'Faridabad', 'Fatehabad', 'Gurugram', 'Hisar', 'Jhajjar', 'Jind', 'Kaithal', 'Karnal', 'Kurukshetra', 'Mahendragarh', 'Mewat', 'Palwal', 'Panchkula', 'Panipat', 'Rewari', 'Rohtak', 'Sirsa', 'Sonipat', 'Yamunanagar'], ['Bilaspur', 'Chamba', 'Hamirpur', 'Kangra', 'Kinnaur', 'Kullu', 'Lahul & Spiti', 'Mandi', 'Shimla', 'Shimla Muncipal Corporation', 'Sirmaur', 'Solan', 'Una'], ['Anantnag', 'Badgam', 'Bandipore', 'Baramula', 'Doda', 'Ganderbal', 'Jammu', 'Kathua', 'kishtwar', 'Kulgam', 'Kupwara', 'Poonch', 'Pulwama', 'Punch', 'Rajauri', 'Ramban', 'Reasi', 'Samba', 'Shopian', 'Srinagar', 'Udhampur'], ['Bokaro', 'Chatra', 'Deoghar', 'Dhanbad', 'Dumka', 'Garhwa', 'Giridih', 'Godda', 'Gumla', 'Hazaribagh', 'Jamtara', 'Khunti', 'Kodarma', 'Latehar', 'Lohardaga', 'Pakaur', 'Palamu', 'Pashchimi Singhbhum', 'Purbi Singhbhum', 'Ramgarh', 'Ranchi', 'Sahibganj', 'Saraikela', 'Simdega'], ['Bagalakote', 'Bangalore Rural', 'Bangalore Urban', 'B.B.M.P East', 'B.B.M.P North', 'B.B.M.P South', 'B.B.M.P West', 'Belgaum', 'Bellary', 'Bidar', 'Bijapur', 'Chamarajanagar', 'Chikkaballapura', 'Chikmagalur', 'Chitradurga', 'Dakshina Kannada', 'Davanagere', 'Dharwad', 'Gadag', 'Gulbarga', 'Hassan', 'Haveri', 'HDMC', 'Kodagu', 'Kolar', 'Koppal', 'Mandya', 'Mysore', 'Raichur', 'Ramanagara', 'Shimoga', 'Tumkur', 'Udupi', 'Uttara Kannada', 'Yadgir'], ['Alappuzha', 'Ernakulam', 'Idukki', 'Kannur', 'Kasaragod', 'Kollam', 'Kottayam', 'Kozhikode', 'Malappuram', 'Palakkad', 'Pathanamthitta', 'Thiruvananthapuram', 'Thrissur', 'Wayanad'], ['Kargil', 'Leh', 'LehLadakh'], ['Agatti', 'Amini', 'Androth', 'Chetlat', 'Kadmath', 'Kalpeni', 'Kavaratti', 'Kiltan', 'Minicoy'], ['Agar Malwa', 'Alirajpur', 'Anuppur', 'Ashoknagar', 'Balaghat', 'Barwani', 'Betul', 'Bhind', 'Bhopal', 'Burhanpur', 'Chhatarpur', 'Chhindwara', 'Damoh', 'Datia', 'Dewas', 'Dhar', 'Dindori', 'Guna', 'Gwalior', 'Harda', 'Indore', 'Jabalpur', 'Jhabua', 'Katni', 'Khandwa', 'Khargone', 'Mandla', 'Mandsaur', 'Morena', 'Narmadapuram', 'Narsimhapur', 'Neemuch', 'Panna', 'Raisen', 'Rajgarh', 'Ratlam', 'Rewa', 'Sagar', 'Satna', 'Sehore', 'Seoni', 'Shahdol', 'Shajapur', 'Sheopur', 'Shivpuri', 'Sidhi', 'Sigrolli', 'Tikamgarh', 'Ujjain', 'Umaria', 'Vidisha'], ['Ahmadnagar', 'Akola', 'Amravati', 'Aurangabad', 'Beed', 'Bhandara', 'Bhiwandi Municipal Corporation (Thane Zone-5)', 'Buldana', 'Chandrapur', 'Dhule', 'Gadchiroli', 'Gondiya', 'Greater Mumbai', 'Hingoli', 'Jalgaon', 'Jalna', 'Kalyan Tahashil (Thane Zone-8)', 'Kolhapur', 'Latur', 'Mira Bhayander Municipal Corporation (Thane Zone-7)', 'Nagpur', 'Nanded', 'Nandurbar', 'Nashik', 'Navi Mumbai Municipal Corporation (Thane Zone-2)', 'Navi Mumbai Municipal Corporation (Thane Zone-3)', 'Navi Mumbai Municipal Corporation (Thane Zone-4)', 'Osmanabad', 'Palghar', 'Parbhani', 'Pune', 'Raigad', 'Ratnagiri', 'Sangli', 'Satara', 'Sindhudurg', 'Solapur', 'Thane Municipal Corporation (Thane Zone-1)', 'Ulhasnagar & Ambarnath Tahashil (Thane Zone-9)', 'Wardha', 'Washim', 'Yavatmal'], ['Bishnupur', 'CHANDEL', 'Churachandpur', 'Imphal East', 'Imphal West', 'JIRIBAM', 'KAKCHING', 'KAMJONG', 'KANGPOKPI', 'NONEY', 'PHERZAWL', 'SENAPATI', 'TAMENGLONG', 'TENGNOUPAL', 'Thoubal', 'UKHRUL'], ['East Garo Hills', 'East Jaintia Hills', 'East Khasi Hills', 'North Garo Hills', 'Ri Bhoi', 'South Garo Hills', 'South West Garo Hills', 'South West Khasi Hills', 'West Garo Hills', 'West Jaintia Hills', 'West Khasi Hills'], ['Aizawl East', 'Aizawl West', 'Champhai', 'Kolasib', 'Lawngtlai', 'Lunglei', 'Mamit', 'Saiha', 'Serchhip'], ['Dimapur', 'KIPHIRE', 'Kohima', 'LONGLENG', 'Mokokchung', 'Mon', 'PEREN', 'Phek', 'Tuensang', 'Wokha', 'Zunheboto'], ['ANGUL', 'BALASORE', 'BARAGARH', 'BERHAMPUR MUNICIPAL CORPORATION', 'BHADRAK', 'BHUBANESWAR MUNICIPAL CORPORATION', 'BOLANGIR', 'BOUDH', 'CUTTACK(EXCEPT MUNICIPAL CORPORATION)', 'CUTTACK MUNICIPAL CORPORATION', 'DEOGARH', 'DHENKANAL', 'GAJAPATI', 'GANJAM(EXCEPT MUNICIPAL CORPORATION)', 'JAGATSINGHPUR', 'JAJPUR', 'JHARSUGUDA', 'KALAHANDI', 'KANDHAMAL', 'KENDRAPARA', 'KEONJHAR', 'KHORDHA(EXCEPT MUNICIPAL CORPORATION)', 'KORAPUT', 'MALKANGIRI', 'MAYURBHANJ', 'NAWARANGPUR', 'NAYAGARH', 'NUAPADA', 'PURI (EXCEPT MUNICIPALITY)', 'PURI MUNICIPALITY', 'RAYAGADA', 'ROURKELA MUNICIPAL CORPORATION', 'SAMBALPUR(EXCEPT MUNICIPAL CORPORATION)', 'SAMBALPUR MUNICIPAL CORPORATION', 'SONEPUR', 'SUNDARGARH(EXCEPT MUNICIPAL CORPORATION)'], ['Puducherry'], ['Amritsar', 'Barnala', 'Bathinda', 'Faridkot', 'Fatehgarh Sahib', 'Fazilka', 'Firozpur', 'Gurdaspur', 'Hoshiarpur', 'Jalandhar', 'Kapurthala', 'Ludhiana', 'Mansa', 'Moga', 'Mohali', 'Muktsar', 'Pathankot', 'Patiala', 'Rupnagar', 'Sangrur', 'Shaheed Bhagat Singh Nagar', 'Taran Taran'], ['Ajmer', 'Alwar', 'Banswara', 'Baran', 'Barmer', 'Bharatpur', 'Bhilwara', 'Bikaner', 'Bundi', 'Chittaurgarh', 'Churu', 'Dausa', 'Dhaulpur', 'Dungarpur', 'Hanumangarh', 'Jaipur', 'Jaisalmer', 'Jalor', 'Jhalawar', 'Jhunjhunun', 'Jodhpur', 'Karauli', 'Kota', 'Nagaur', 'Pali', 'Pratapgarh', 'Rajsamand', 'Sawai Madhopur', 'Shri Ganganagar', 'Sikar', 'Sirohi', 'Tonk', 'Udaipur'], ['East', 'Gangtok', 'North', 'South', 'West'], ['Ariyalur', 'Chennai', 'Coimbatore', 'Cuddalore', 'Dharmapuri', 'Dindigul', 'Erode', 'Kancheepuram', 'Kanyakumari', 'Karur', 'Krishnagiri', 'Madurai', 'Nagapattinam', 'Namakkal', 'Perambalur', 'Pudukkottai', 'Ramanathapuram', 'Salem', 'Sivaganga', 'Thanjavur', 'Theni', 'The Nilgiris', 'Thiruvallur', 'Thiruvarur', 'Thoothukkudi', 'Tiruchirappalli', 'Tirunelveli', 'Tirupur', 'Tiruvannamalai', 'Vellore', 'Viluppuram', 'Virudhunagar'], ['Adilabad', 'Badradri - Kothagudem', 'Cantonment-Sec-Bad', 'Hyderabad', 'Jagitial', 'Jangoan', 'Jayashanker-Bhoopalpally', 'Jogulamba - Gadwal', 'Kamareddy', 'Karimnagar', 'Khammam', 'Komarambheem', 'Mahabubabad', 'Mahbubnagar', 'Mancherial', 'Medak', 'Medchal-Malkajgiri', 'Nagar Kurnool', 'Nalgonda', 'Nirmal', 'Nizamabad', 'Peddapally', 'Rajanna', 'Rangareddi', 'Sangareddy', 'Siddipet', 'Suryapet', 'Vikarabad', 'Wanaparthy', 'Warangal', 'Yadadri - Bhongir'], ['Agartala Municipal Council', 'Dhalai', 'Gomati', 'Khowai', 'North Tripura', 'Sepahijala', 'South Tripura', 'Unakoti', 'West Tripura'], ['Almora', 'Bageshwar', 'Chamoli', 'Champawat', 'Dehradun', 'Hardwar', 'Nainital', 'Pauri Garhwal', 'Pithoragarh', 'Rudraprayag', 'Tehri Garhwal', 'Udham Singh Nagar', 'Uttarkashi'], ['Agra', 'Aligarh', 'Ambedkar Nagar', 'Amethi', 'Amroha', 'Auraiya', 'Ayodhya', 'Azamgarh', 'Baghpat', 'Bahraich', 'Ballia', 'Balrampur', 'Banda', 'Barabanki', 'Bareilly', 'Basti', 'Bijnor', 'Budaun', 'Bulandshahar', 'Chandauli', 'Chitrakoot', 'Deoria', 'Etah', 'Etawah', 'Farrukhabad', 'Fatehpur', 'Firozabad', 'Gautam Buddha Nagar', 'Ghaziabad', 'Ghazipur', 'Gonda', 'Gorakhpur', 'Hamirpur', 'Hapur', 'Hardoi', 'Hathras', 'Jalaun', 'Jaunpur', 'Jhansi', 'Kannauj', 'Kanpur Dehat', 'Kanpur Nagar', 'Kasganj', 'Kaushambi', 'Kheri', 'Kushinagar', 'Lakhimpur Khiri', 'Lalitpur', 'Lucknow', 'Mahoba', 'Mahrajganj', 'Mainpuri', 'Mathura', 'Mau', 'Meerut', 'Mirzapur', 'Moradabad', 'Muzaffarnagar', 'Pilibhit', 'Pratapgarh', 'Prayagraj', 'Rae Bareli', 'Rampur', 'Saharanpur', 'Sambhal', 'Sant Kabir Nagar', 'Sant Ravidas Nagar Bhadohi', 'Shahjahanpur', 'Shamali', 'Shrawasti', 'Siddharthnagar', 'Sitapur', 'Sonbhadra', 'Sultanpur', 'Unnao', 'Varanasi'], ['Alipurduar', 'Bankura', 'Basirhat Health District', 'Birbhum', 'Bishnupur Health District', 'Coochbehar', 'Dakshin Dinajpur', 'Darjeeling', 'DIAMOND HARBOUR HEALTH DISTRICT', 'Hooghly', 'Howrah', 'Jalpaiguri', 'JHARGRAM', 'Kalimpong', 'Kolkata', 'Maldah', 'Murshidabad', 'Nadia', 'NANDIGRAM HEALTH DISTRICT', 'North Twenty Four Parganas', 'PASCHIM BARDHAMAN', 'Paschim Mednipur', 'PURBA BARDHAMAN', 'Purba Medinipur', 'Puruliya', 'Rampurhat Health District', 'South Twenty Four Parganas', 'Uttar Dinajpur']]

#try:
while True:
    print("waiting")
    #try:
    unReadMsg()
    # All the unread Msg is cleared ... So create a Excel file and check for each contact
    query = lastMsg()
    textAndReply(query)
    CloseChat()
    #except:
    #    print("Errors")
#except:
#    print("ReRun\nKindly Check Buddy")
