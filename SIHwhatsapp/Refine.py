import sys
import time
import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

def unReadMsg():
    unReadMsg = WebDriverWait(browser, 1000).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'div[class="_1pJ9J"]')))
    unReadMsg.click()


def lastMsg():
    time.sleep(5)
    try:
        LastMsg = browser.find_elements(By.CSS_SELECTOR, 'div[class="_2wUmf message-in focusable-list-item"]')[-1].text.split('\n')[0]
    except:
        print("LeftClick Error ... Maybe")
        browser.refresh()
    # This LastMsg returs 2 string seperated by \n so 1st line has actual data and 2nd line with timestamp
    myLastMsg = browser.find_elements(By.CSS_SELECTOR, 'div[class="_2wUmf message-out focusable-list-item"]')#[-1].text.split('\n')[0]
    if myLastMsg != []:
        myLastMsg = myLastMsg[-1].text.split('\n')[0]
    print(myLastMsg)
    return [LastMsg, myLastMsg]


def textAndReply(query):
    TBox = browser.find_element(By.CSS_SELECTOR, 'p[class="selectable-text copyable-text"]')
    if query[0]=='bye':
        sys.exit()
    if query[1] == []:
        TBox.send_keys(intro)
        return 0
    if 'S' in query[1]:
        pass#Qdata = SProcessor(query)
    if 'Q' in query[1]:
        data = QProcessor(query)
    if 'D' in query[1]:
        pass#data = DProcessor(query)
    TBox.send_keys(data)

def CloseChat():
    browser.find_element(By.CSS_SELECTOR,'div[data-tab="6"][title="Menu"]').click()
    browser.find_element(By.CSS_SELECTOR,'div[aria-label="Close chat"]').click()

def PressFormate(part):
    partion=''
    for i in range(0, len(part)):
        partion+="Press "+str( (i+1))+ " for "+str(part[i])+ "\n"
    return partion

def getDataFromSheet():
    contacts=[]
    for i in range(1,sheet.max_row+1):
        contacts.append(sheet["A"+str(i)].value)
    contactNo =browser.find_element(By.CSS_SELECTOR,'div[class="_24-Ff"]').text.split("\n")
    if contactNo[0] not in contacts:
        print("Appending")
        row=sheet.max_row+1
        sheet["A"+str(row)].value = contactNo[0]
        workspace.save("DetailsStorage.xlsx")
        return row
    else:
        pos = contacts.index(contactNo[0])+1
        return pos


def fileGet(code):
    path = r"SubDist/D" + code + ".txt"
    f = open(path,"r")
    content = f.read()
    script = "global l\nl = " + content
    exec(script)
    return l
def getSubDistricts(query):
    code = query[1].replace("S","")
    content = fileGet(code)
    return (content[int(query[0])-1])
def getSubDist(query):
    code = query[1].replace("S","")
    code = code.split("_")
    content = fileGet(code[0])
    dist = content[int(code[1])-1]
    return dist[int(query[0])-1]

def stepExcelInclude(columbLetter,query):
    cell = getDataFromSheet()
    print(cell, query, columbLetter + str(cell))
    write = sheet[columbLetter + str(cell)]
    write.value = query
    workspace.save(r"DetailsStorage.xlsx")




def QProcessor(query):
    if query[1] == "Q0":
        if query[0].isnumeric():
            if int(query[0])>=1 and int(query[0]<=3):
                return "You have selected %s\n%s\nQ1\n"%("File a Complaint",PressFormate(catagory))
            else:
                return "Invalid.\n"+intro
        else:
            return "Invalid.\n" + intro
    if query[1] == "Q1":
        columnLetter="B"
        if query[0].isnumeric():
            if query[0]>=1 and query[0]<=5:
                stepExcelInclude(columnLetter,catagory[int(query[0])-1])
                return "%s\n%s\n"%(subSelect[int(query[0])-1],"Q"+query[0])
            else:
                return "Invalid\n%s\nQ1"%PressFormate(catagory)
        else:
            return "Invalid\n%s\nQ1"%PressFormate(catagory)
    if query[1]:
        columnLetter = "C"
        code = query[1]
        code = code.replace("Q","")
        code = int(code[-1])-1
        if query[0].isnumeric():
            if int(query[0]) >= 1 and int(query[0]) <= len(subSelect[code]):
                stepExcelInclude(columnLetter, subSelect[int(query[0])-1])
                return "SubCategory selected\nEnter the Product Name:\nM1\n"
            else:
                return "Invalid\n%s\n%s"%(subSelect[code],query[1])
        else:
            return "Invalid\n%s\n%s"%(subSelect[code],query[1])
    workspace.save("DetailsStorage.xlsx")

# ---------------------------------------------------------------------------------------#
intro="Hi\nThis is a place where we help You to file a case in fssai website\nPress 1 to File a case\nPress 2 to track the case\nPress 3 to know your rights\nQ0\n"
catagory = ['Package Food', 'Food Catering Premises', 'Online aggregator/e-commerce', 'Retailer Premises', 'Others']
subCatOfPack = ['Dairy products', 'Fats & oils', 'Edible ices including sorbet', 'Confectionery', 'Cereal & cereal products', 'Bakery products', 'Meat & meat products including poultry', 'Fish & fish products', 'Egg & egg products', 'Sweetners including honey', 'Salt, spices, soups, sauces, salads & protein products', 'Beverages excluding dairy products', 'Ready to eat savouries', 'Prepared food', 'Others', 'Nutraceuticals', 'Fruit and Vegetables']
subCatOfCate = ['Restaurants ', 'Canteen ', 'Cafeteria ', 'Dhabas ', 'Cafe ', 'Hostel Mess ', 'Food Trucks ', 'Take aways ', 'Hotel ', 'Others ']
subCatOfOn = ['Prepared Food Delivering Agency ', 'Grocery Delivering Agency ', 'Others']
subCatOfRet = ['Retail shops ', 'Milk & milk products retail shop ', 'Meat & meat products (including poultry & fish) retail shop ', 'Fruits & vegetable retail shop ', 'Others ']
subCatOfOth = ["Others"]
subSelect = [subCatOfPack, subCatOfCate, subCatOfOn, subCatOfRet, subCatOfOth]
State=['Andaman And Nicobar Islands', 'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chandigarh', 'Chhattisgarh', 'Dadra & Nagar Haveli', 'Daman & Diu', 'Delhi', 'Goa', 'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jammu & Kashmir', 'Jharkhand', 'Karnataka', 'Kerala', 'Ladakh', 'Lakshadweep', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya', 'Mizoram', 'Nagaland', 'Orissa', 'Puducherry', 'Punjab', 'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana', 'Tripura', 'Uttarakhand', 'Uttar Pradesh', 'West Bengal']
listOfDist = [['Andamans', 'Nicobars'], ['Anantapur', 'Chittoor', 'Cuddapah', 'East Godavari', 'Guntur', 'Kadapa Municipal Corporation', 'Kakinada Mpl.Corp.', 'Krishna', 'Kurnool', 'Nellore', 'Prakasam', 'Srikakulam', 'Visakhapatnam', 'Vizianagaram', 'West Godavari'], ['Anjaw', 'Changlang', 'Dibang Valley', 'East Kameng', 'East Siang', 'Kamle', 'Kra Dadi', 'Kurung Kumey', 'Leparada', 'Lohit', 'Longding', 'Lower Dibang Valley', 'Lower Subansiri', 'Lower Subansiri', 'Namsai', 'Pakke Kessang', 'Papum Pare', 'Shi Yomi', 'Siang', 'Tawang', 'Tirap', 'Upper Siang', 'Upper Subansiri', 'West Kameng', 'West Siang'], ['Baksa', 'Barpeta', 'Biswanath', 'Bongaigaon', 'Cachar', 'Charaideo', 'Chirang', 'Darrang', 'Dhemaji', 'Dhubri', 'Dibrugarh', 'Dima Hasao', 'Goalpara', 'Golaghat', 'Hailakandi', 'Hojai', 'Jorhat', 'Kamrup Dist.', 'Kamrup Metropolitan Dist.', 'Karbi Anglong', 'Karimganj', 'Kokrajhar', 'Lakhimpur', 'Majuli', 'Marigaon', 'Nagaon', 'Nalbari', 'Sivasagar', 'Sonitpur', 'South Salmara Mancachar', 'Tinsukia', 'Udalguri', 'West Karbi Anglong'], ['Araria', 'Arwal', 'Aurangabad', 'Banka', 'Begusarai', 'Bhagalpur', 'Bhojpur', 'Buxar', 'Darbhanga', 'Gaya', 'Gopalganj', 'Jamui', 'Jehanabad', 'Kaimur (Bhabua)', 'Katihar', 'Khagaria', 'Kishanganj', 'Lakhisarai', 'Madhepura', 'Madhubani', 'Munger', 'Muzaffarpur', 'Nalanda', 'Nawada', 'Pashchim Champaran', 'Patna', 'Purba Champaran', 'Purnia', 'Rohtas', 'Saharsa', 'Samastipur', 'Saran', 'Sheikhpura', 'Sheohar', 'Sitamarhi', 'Siwan', 'Supaul', 'Vaishali'], ['Chandigarh'], ['Balod', 'Balodabazar', 'Balrampur', 'Bastar', 'Bemetara', 'Bijapur', 'Bilaspur', 'Dantewada', 'Dhamtari', 'Durg', 'Gariyaband', 'Gurela-Pendra-Marwahi', 'Janjgir - Champa', 'Jashpur', 'Kabirdham', 'Kanker', 'Kawardha', 'Kondagaon', 'Korba', 'KOREA', 'Mahasamund', 'Mungeli', 'Narayanpur', 'Raigarh', 'Raipur', 'Rajnandgaon', 'Sukma', 'Surajpur', 'Surguja'], ['Dadra & Nagar Haveli'], ['Daman', 'Diu'], ['Central', 'East', 'New Delhi', 'North', 'North East', 'North West', 'Shahdara', 'South', 'South East', 'South West', 'West'], ['North Goa', 'South Goa'], ['Ahmedabad', 'Amreli', 'Amreli Municipality', 'Anand', 'Arvalli', 'BANASKANTHA', 'Bharuch', 'Bhavnagar', 'BHUJ(KUTCHH)', 'Botad Rural', 'Chhota Udaipur', 'Dahod Municipality', 'Dang', 'Devbhumi Dwarka', 'Dohad', 'Gandhidham Municipality', 'Gandhinagar', 'Gir Somnath', 'Godhra Municipality', 'GODHRA(PANCHMAHAL)', 'HIMMATNAGAR(SABARKANTHA)', 'Jamnagar', 'Junagadh', 'Kalol Municipality', 'Mahesana', 'MAHISAGAR', 'MORBI', 'NADIAD(KHEDA)', 'Narmada', 'Navsari', 'Patan', 'Porbandar', 'Rajkot', 'Rajkot Municipal Corporation', 'Surat', 'Surat Municipal Corporation', 'Surendranagar', 'Tapi', 'Vadodara', 'Vadodara Municipal Corporation', 'Valsad', 'VYARA(TAPI)'], ['Ambala', 'Bhiwani', 'Charkhi Dadri', 'Faridabad', 'Fatehabad', 'Gurugram', 'Hisar', 'Jhajjar', 'Jind', 'Kaithal', 'Karnal', 'Kurukshetra', 'Mahendragarh', 'Mewat', 'Palwal', 'Panchkula', 'Panipat', 'Rewari', 'Rohtak', 'Sirsa', 'Sonipat', 'Yamunanagar'], ['Bilaspur', 'Chamba', 'Hamirpur', 'Kangra', 'Kinnaur', 'Kullu', 'Lahul & Spiti', 'Mandi', 'Shimla', 'Shimla Muncipal Corporation', 'Sirmaur', 'Solan', 'Una'], ['Anantnag', 'Badgam', 'Bandipore', 'Baramula', 'Doda', 'Ganderbal', 'Jammu', 'Kathua', 'kishtwar', 'Kulgam', 'Kupwara', 'Poonch', 'Pulwama', 'Punch', 'Rajauri', 'Ramban', 'Reasi', 'Samba', 'Shopian', 'Srinagar', 'Udhampur'], ['Bokaro', 'Chatra', 'Deoghar', 'Dhanbad', 'Dumka', 'Garhwa', 'Giridih', 'Godda', 'Gumla', 'Hazaribagh', 'Jamtara', 'Khunti', 'Kodarma', 'Latehar', 'Lohardaga', 'Pakaur', 'Palamu', 'Pashchimi Singhbhum', 'Purbi Singhbhum', 'Ramgarh', 'Ranchi', 'Sahibganj', 'Saraikela', 'Simdega'], ['Bagalakote', 'Bangalore Rural', 'Bangalore Urban', 'B.B.M.P East', 'B.B.M.P North', 'B.B.M.P South', 'B.B.M.P West', 'Belgaum', 'Bellary', 'Bidar', 'Bijapur', 'Chamarajanagar', 'Chikkaballapura', 'Chikmagalur', 'Chitradurga', 'Dakshina Kannada', 'Davanagere', 'Dharwad', 'Gadag', 'Gulbarga', 'Hassan', 'Haveri', 'HDMC', 'Kodagu', 'Kolar', 'Koppal', 'Mandya', 'Mysore', 'Raichur', 'Ramanagara', 'Shimoga', 'Tumkur', 'Udupi', 'Uttara Kannada', 'Yadgir'], ['Alappuzha', 'Ernakulam', 'Idukki', 'Kannur', 'Kasaragod', 'Kollam', 'Kottayam', 'Kozhikode', 'Malappuram', 'Palakkad', 'Pathanamthitta', 'Thiruvananthapuram', 'Thrissur', 'Wayanad'], ['Kargil', 'Leh', 'LehLadakh'], ['Agatti', 'Amini', 'Androth', 'Chetlat', 'Kadmath', 'Kalpeni', 'Kavaratti', 'Kiltan', 'Minicoy'], ['Agar Malwa', 'Alirajpur', 'Anuppur', 'Ashoknagar', 'Balaghat', 'Barwani', 'Betul', 'Bhind', 'Bhopal', 'Burhanpur', 'Chhatarpur', 'Chhindwara', 'Damoh', 'Datia', 'Dewas', 'Dhar', 'Dindori', 'Guna', 'Gwalior', 'Harda', 'Indore', 'Jabalpur', 'Jhabua', 'Katni', 'Khandwa', 'Khargone', 'Mandla', 'Mandsaur', 'Morena', 'Narmadapuram', 'Narsimhapur', 'Neemuch', 'Panna', 'Raisen', 'Rajgarh', 'Ratlam', 'Rewa', 'Sagar', 'Satna', 'Sehore', 'Seoni', 'Shahdol', 'Shajapur', 'Sheopur', 'Shivpuri', 'Sidhi', 'Sigrolli', 'Tikamgarh', 'Ujjain', 'Umaria', 'Vidisha'], ['Ahmadnagar', 'Akola', 'Amravati', 'Aurangabad', 'Beed', 'Bhandara', 'Bhiwandi Municipal Corporation (Thane Zone-5)', 'Buldana', 'Chandrapur', 'Dhule', 'Gadchiroli', 'Gondiya', 'Greater Mumbai', 'Hingoli', 'Jalgaon', 'Jalna', 'Kalyan Tahashil (Thane Zone-8)', 'Kolhapur', 'Latur', 'Mira Bhayander Municipal Corporation (Thane Zone-7)', 'Nagpur', 'Nanded', 'Nandurbar', 'Nashik', 'Navi Mumbai Municipal Corporation (Thane Zone-2)', 'Navi Mumbai Municipal Corporation (Thane Zone-3)', 'Navi Mumbai Municipal Corporation (Thane Zone-4)', 'Osmanabad', 'Palghar', 'Parbhani', 'Pune', 'Raigad', 'Ratnagiri', 'Sangli', 'Satara', 'Sindhudurg', 'Solapur', 'Thane Municipal Corporation (Thane Zone-1)', 'Ulhasnagar & Ambarnath Tahashil (Thane Zone-9)', 'Wardha', 'Washim', 'Yavatmal'], ['Bishnupur', 'CHANDEL', 'Churachandpur', 'Imphal East', 'Imphal West', 'JIRIBAM', 'KAKCHING', 'KAMJONG', 'KANGPOKPI', 'NONEY', 'PHERZAWL', 'SENAPATI', 'TAMENGLONG', 'TENGNOUPAL', 'Thoubal', 'UKHRUL'], ['East Garo Hills', 'East Jaintia Hills', 'East Khasi Hills', 'North Garo Hills', 'Ri Bhoi', 'South Garo Hills', 'South West Garo Hills', 'South West Khasi Hills', 'West Garo Hills', 'West Jaintia Hills', 'West Khasi Hills'], ['Aizawl East', 'Aizawl West', 'Champhai', 'Kolasib', 'Lawngtlai', 'Lunglei', 'Mamit', 'Saiha', 'Serchhip'], ['Dimapur', 'KIPHIRE', 'Kohima', 'LONGLENG', 'Mokokchung', 'Mon', 'PEREN', 'Phek', 'Tuensang', 'Wokha', 'Zunheboto'], ['ANGUL', 'BALASORE', 'BARAGARH', 'BERHAMPUR MUNICIPAL CORPORATION', 'BHADRAK', 'BHUBANESWAR MUNICIPAL CORPORATION', 'BOLANGIR', 'BOUDH', 'CUTTACK(EXCEPT MUNICIPAL CORPORATION)', 'CUTTACK MUNICIPAL CORPORATION', 'DEOGARH', 'DHENKANAL', 'GAJAPATI', 'GANJAM(EXCEPT MUNICIPAL CORPORATION)', 'JAGATSINGHPUR', 'JAJPUR', 'JHARSUGUDA', 'KALAHANDI', 'KANDHAMAL', 'KENDRAPARA', 'KEONJHAR', 'KHORDHA(EXCEPT MUNICIPAL CORPORATION)', 'KORAPUT', 'MALKANGIRI', 'MAYURBHANJ', 'NAWARANGPUR', 'NAYAGARH', 'NUAPADA', 'PURI (EXCEPT MUNICIPALITY)', 'PURI MUNICIPALITY', 'RAYAGADA', 'ROURKELA MUNICIPAL CORPORATION', 'SAMBALPUR(EXCEPT MUNICIPAL CORPORATION)', 'SAMBALPUR MUNICIPAL CORPORATION', 'SONEPUR', 'SUNDARGARH(EXCEPT MUNICIPAL CORPORATION)'], ['Puducherry'], ['Amritsar', 'Barnala', 'Bathinda', 'Faridkot', 'Fatehgarh Sahib', 'Fazilka', 'Firozpur', 'Gurdaspur', 'Hoshiarpur', 'Jalandhar', 'Kapurthala', 'Ludhiana', 'Mansa', 'Moga', 'Mohali', 'Muktsar', 'Pathankot', 'Patiala', 'Rupnagar', 'Sangrur', 'Shaheed Bhagat Singh Nagar', 'Taran Taran'], ['Ajmer', 'Alwar', 'Banswara', 'Baran', 'Barmer', 'Bharatpur', 'Bhilwara', 'Bikaner', 'Bundi', 'Chittaurgarh', 'Churu', 'Dausa', 'Dhaulpur', 'Dungarpur', 'Hanumangarh', 'Jaipur', 'Jaisalmer', 'Jalor', 'Jhalawar', 'Jhunjhunun', 'Jodhpur', 'Karauli', 'Kota', 'Nagaur', 'Pali', 'Pratapgarh', 'Rajsamand', 'Sawai Madhopur', 'Shri Ganganagar', 'Sikar', 'Sirohi', 'Tonk', 'Udaipur'], ['East', 'Gangtok', 'North', 'South', 'West'], ['Ariyalur', 'Chennai', 'Coimbatore', 'Cuddalore', 'Dharmapuri', 'Dindigul', 'Erode', 'Kancheepuram', 'Kanyakumari', 'Karur', 'Krishnagiri', 'Madurai', 'Nagapattinam', 'Namakkal', 'Perambalur', 'Pudukkottai', 'Ramanathapuram', 'Salem', 'Sivaganga', 'Thanjavur', 'Theni', 'The Nilgiris', 'Thiruvallur', 'Thiruvarur', 'Thoothukkudi', 'Tiruchirappalli', 'Tirunelveli', 'Tirupur', 'Tiruvannamalai', 'Vellore', 'Viluppuram', 'Virudhunagar'], ['Adilabad', 'Badradri - Kothagudem', 'Cantonment-Sec-Bad', 'Hyderabad', 'Jagitial', 'Jangoan', 'Jayashanker-Bhoopalpally', 'Jogulamba - Gadwal', 'Kamareddy', 'Karimnagar', 'Khammam', 'Komarambheem', 'Mahabubabad', 'Mahbubnagar', 'Mancherial', 'Medak', 'Medchal-Malkajgiri', 'Nagar Kurnool', 'Nalgonda', 'Nirmal', 'Nizamabad', 'Peddapally', 'Rajanna', 'Rangareddi', 'Sangareddy', 'Siddipet', 'Suryapet', 'Vikarabad', 'Wanaparthy', 'Warangal', 'Yadadri - Bhongir'], ['Agartala Municipal Council', 'Dhalai', 'Gomati', 'Khowai', 'North Tripura', 'Sepahijala', 'South Tripura', 'Unakoti', 'West Tripura'], ['Almora', 'Bageshwar', 'Chamoli', 'Champawat', 'Dehradun', 'Hardwar', 'Nainital', 'Pauri Garhwal', 'Pithoragarh', 'Rudraprayag', 'Tehri Garhwal', 'Udham Singh Nagar', 'Uttarkashi'], ['Agra', 'Aligarh', 'Ambedkar Nagar', 'Amethi', 'Amroha', 'Auraiya', 'Ayodhya', 'Azamgarh', 'Baghpat', 'Bahraich', 'Ballia', 'Balrampur', 'Banda', 'Barabanki', 'Bareilly', 'Basti', 'Bijnor', 'Budaun', 'Bulandshahar', 'Chandauli', 'Chitrakoot', 'Deoria', 'Etah', 'Etawah', 'Farrukhabad', 'Fatehpur', 'Firozabad', 'Gautam Buddha Nagar', 'Ghaziabad', 'Ghazipur', 'Gonda', 'Gorakhpur', 'Hamirpur', 'Hapur', 'Hardoi', 'Hathras', 'Jalaun', 'Jaunpur', 'Jhansi', 'Kannauj', 'Kanpur Dehat', 'Kanpur Nagar', 'Kasganj', 'Kaushambi', 'Kheri', 'Kushinagar', 'Lakhimpur Khiri', 'Lalitpur', 'Lucknow', 'Mahoba', 'Mahrajganj', 'Mainpuri', 'Mathura', 'Mau', 'Meerut', 'Mirzapur', 'Moradabad', 'Muzaffarnagar', 'Pilibhit', 'Pratapgarh', 'Prayagraj', 'Rae Bareli', 'Rampur', 'Saharanpur', 'Sambhal', 'Sant Kabir Nagar', 'Sant Ravidas Nagar Bhadohi', 'Shahjahanpur', 'Shamali', 'Shrawasti', 'Siddharthnagar', 'Sitapur', 'Sonbhadra', 'Sultanpur', 'Unnao', 'Varanasi'], ['Alipurduar', 'Bankura', 'Basirhat Health District', 'Birbhum', 'Bishnupur Health District', 'Coochbehar', 'Dakshin Dinajpur', 'Darjeeling', 'DIAMOND HARBOUR HEALTH DISTRICT', 'Hooghly', 'Howrah', 'Jalpaiguri', 'JHARGRAM', 'Kalimpong', 'Kolkata', 'Maldah', 'Murshidabad', 'Nadia', 'NANDIGRAM HEALTH DISTRICT', 'North Twenty Four Parganas', 'PASCHIM BARDHAMAN', 'Paschim Mednipur', 'PURBA BARDHAMAN', 'Purba Medinipur', 'Puruliya', 'Rampurhat Health District', 'South Twenty Four Parganas', 'Uttar Dinajpur']]

l = []
workspace = openpyxl.load_workbook(r"DetailsStorage.xlsx")
sheet = workspace.active

opt = Options()
#options.add_experimental_option("debuggerAddress","localhost:1234")
opt.add_argument(r'--user-data-dir=E:\Hackathon\BrowserChromes\Whatsapp')
#opt.add_argument("--headless")

services = Service(r"C:\Users\HP\PycharmProjects\WebDriver\chromedriver.exe")

browser = Chrome(service=services,options=opt)

browser.get("https://web.whatsapp.com/")
print("Opened")
browser.implicitly_wait(15)

while True:
    print("waiting")
    #try:
    unReadMsg()
    # All the unread Msg is cleared ... So create a Excel file and check for each contact
    query = lastMsg()
    textAndReply(query)
    CloseChat()
