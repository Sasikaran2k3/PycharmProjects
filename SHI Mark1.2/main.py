import sys
import time
import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By


def unReadMsg():
    unReadMsg = WebDriverWait(browser, 1000).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'div[class="_1pJ9J"]')))
    unReadMsg.click()

def getRowOfContact():
    contacts=[]
    for i in range(1,sheet.max_row+1):
        contacts.append(sheet["A"+str(i)].value)
    contactNo = browser.find_element(By.CSS_SELECTOR,'div[class="_24-Ff"]').text.split("\n")
    if contactNo[0] not in contacts:
        print("Appending")
        row=sheet.max_row+1
        sheet["A"+str(row)].value = contactNo[0]
        workspace.save("DetailsStorage.xlsx")
        return row
    else:
        pos = contacts.index(contactNo[0])+1
        return pos

def getStep(contactNo):
    step=0
    i=1
    while sheet.cell(contactNo,i).value != None:
        step+=1
        i+=1
    prev = seq[step-1]
    unFill = seq[step]
    return [prev,unFill]
def FillForm(proStep):
    if proStep[1] == "State":


seq = ["Contact", "Category", "SubCategory", "State", "District", "Sub-District"]
workspace = openpyxl.load_workbook(r"DetailsStorage.xlsx")
sheet = workspace.active
opt=Options()
opt.add_argument(r'--user-data-dir=E:\Chatterbot\Selenium Browser')
#opt.add_argument("--headless")
services = Service(r"C:\Users\HP\PycharmProjects\WebDriver\chromedriver.exe")
browser=Chrome(service=services,options=opt)
browser.get("https://web.whatsapp.com/")
browser.implicitly_wait(15)
print("Opened")


unReadMsg()
proStep = getStep(getRowOfContact())
