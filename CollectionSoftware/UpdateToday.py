import datetime
import openpyxl
from openpyxl import load_workbook, Workbook

date = "".join(str(datetime.date.today()).split("-"))

wb = load_workbook(date+".xlsx")
pointer = wb.active

pointer.delete_rows(1,2)
wb.save(date+".xlsx")

list_of_shop = pointer["C"]
list_of_amt = pointer["D"]

for i in range(len(list_of_shop)):
    try:
        add_wb = load_workbook(list_of_shop[i].value+".xlsx")
        add_pointer = add_wb.active
        add_pointer.append((str(datetime.date.today()),list_of_amt[i].value))
        add_wb.save(list_of_shop[i].value+".xlsx")
    except:
        pass