import datetime
import os.path
from tkinter import *
from tkinter.ttk import Combobox

import openpyxl
from openpyxl import load_workbook

root = Tk()
root.title("Ravi Oil Store")
img = PhotoImage(file='ros_logo.png')
root.iconphoto(False, img)
root.geometry("1080x720")
area_list = ["Full","Choolaimedu", "Anna Nagar", "MMDA","Arumbakkam"]
Choolaimedu = ["ARI ALAGAR STORES", "ARIHARAN STORES", "ARTHI FISH STORE", "ARTHI STORE(ARUM)"]
anna_nagar = ['AYYANAR STORE (Y-20)', 'CHANDRA STORE', 'SRI RAMAJAYAM STORE', 'PADMA STORE', 'DHANALAKSHMI STORE', 'SIVA STORE (THANGA COLONY)', 'CHANDRA SUPER MARKET', 'MURUGAN STORE (18TH)', 'INDIRA STORE (11 TH)']
mmda = ["RENGA STORE", "THIRUPATHI STORE", "PATHUBANATHAN STORE", "ISHWARYA SWEETS", "SELVA GANESH STORE", "KAVITHA STORE", "AK SUPER MARKET", "MANGALAM SUPER MARKET", "KAMATCHI STORE (MMDA)"]
arumbakkam = ['SELVAM STORE(ARUM)', 'SRI GANESH STORE (JAI)', 'JAYAMANI STORE', 'LAKSHMAN & LAKSHMI SWEETS', 'SANTHOSH STORE(ARUM)', 'ARTHI STORE(ARUM)']
l = ["ARI ALAGAR STORES", "ARIHARAN STORES", "ARTHI FISH STORE", "ARTHI STORE(ARUM)", "ARUL HOTEL", "ARUL STORE (CHOO)", "ARUL SUPER STORE", "ARUN STORES", "ARUN STORES(KOYM)", "ARUNACHALA TRADERS", "ASHVINI PROVISION", "ASHWIN PALAMUDIR NILAYAM", "ASR VELUS PAZHAMUDIR CHOLAI & SM (NOLAM)", "ASR VELUS PAZHAMUDIR CHOLAI & SUPER MARKET", "AST PAZHAMUDHIR NILAYAM", "ATHITHI STORE", "AVG REDDY", "AVP CHANDRA STORE", "AVS STORE", "AYYANAR OIL STORE", "AYYANAR STOER (K-87)", "AYYANAR STORE (G-87)", "AYYANAR STORE (GA307)", "AYYANAR STORE (KIL)", "AYYANAR STORE (THIRU)", "AYYANAR STORE (Y-20)", "AYYANAR STORE NSK", "AYYANAR STORE(CHOOLAI)", "AYYANAR STORE(NAVEL)", "AYYANAR TRADERS(GB-74)", "AYYAPPA STORE (ANNA NAGAR)", "AYYAPPA STORE(SHENOY)", "AYYNAR STORE (NSK 1)", "BABU STORE (ARUM)", "BAGAVATHI AMMAN FRUITS", "BAKYAM STORE (CMBT)", "BAKYAM TRADERS(B-79)", "BALAJI OIL STORE", "BALAJI STORE(AYAN)", "BALAJI STORE(KOYM)", "SRI BALAMURUGAN STORE", "no use BALAMURUGAN STORE(ANNA)", "BARANI STORE", "BARANI STORE(59)", "BHARANI BHAVAN", "BHARANI FAST FOODS", "BISMILLAH STORE", "BJ ENTERPRISES", "BKR STORE", "BR AGENCY", "CHANDRA STORE", "CHANDRA STORE MMDA", "CHANDRA STORE(KOYM)", "CHANDRA SUPER MARKET", "CHENNAI FOODS", "CHENNAI HERITAGE HOSPITAL", "CHENNAI SUPER STORE", "CHENTHUR MURUGAN TRADERS", "CHIDAMBARAM STORE", "CHOICE PROVISION STORE", "COESS VENTURES", "D.S.NAIDU STORE", "DAILY DELITE SUPER MARKET", "DAILY NEEDS", "DEEPAM FOODS(THIRU)", "DEIVAM EVERGREEN SUPERMARKET PVT LTD", "DEVAR PAZHAMUDHIR NILAYAM", "DEVI KARUMARI STORE", "DHANALAKSHMI STORE", "DHANASEI", "DINDUKAL BIRYANI MESS", "DIVIYA STORE", "DIVYA TRADERS", "DURGA (NATTU)", "ESSENTIAL4U", "ESSENTIAL4U(ANNA)", "ESWARI OIL STORES", "EVERSHINE ENTERPRISE", "EVERY DAILY HOME NEEDS", "FRESH 2 DAY", "FRESH OF FINE GERERAL STORE", "FRESH SUPERMARKET", "GANAPATHI BHAVAN", "GANDHIRAJAN STORE (GSK)", "GANESH BHAVAN HOTEL(VP)", "GANESH STORE (NEW AVD RD)", "GANESH STORE(13TH)", "GANESH STORE(LOTUS)", "GANESH STORE(N.S.K.)", "GANESH STORE(SHENOY)", "GANGA SUPER STORE", "GANGAS TIFFEN CENTRE", "GANI STORE", "GIRIJAPAATI FOODS PVT LTD", "GNANAM STORE", "GOOD LIFE ENTERPRISES", "GOOD LIFE SUPER MARKET", "GOOD LUCK STORE", "GOODLIFE SUPERMART (ANNA)", "GR PROVISION", "GREEEN LAND BAKERY", "GREEN MARKETING", "GROCERMAIL E-COMMERCE VENTURES PRIVATE LIMITED", "GSK SUPER MARKET", "HANEEFA SUPER MARKET", "HARI KRISHNA STORE", "HARI OM PHARMA", "HARICHANDRA TRADERS", "HARIYANTH PROVISION", "HOME NEEDS", "HOT CHIPS & SNACKS", "HOT CHIPS & SNACKS (NEW)", "HOT CHIPS (KIL)", "HOT CHIPS(ANNA)", "HOT CHIPS(CHOO)", "HOTEL GANAPATHY BHAVAN", "HOTEL SRI RAJ BHAVAN", "IBRAHIM STORE", "INDIRA STORE (11 TH)", "INDIRA STORE(8TH)", "ISHWARYA SWEETS", "ISMAIL STORE (NELSON)", "ISMAIL STORE(KIL)", "JAFFER STORE", "JAI AKASH STORE", "JAI DURGA STORE", "JAMAILA STORE", "JAMAL MOIDEEN STORE", "JASMINE SUPER MARKET", "JAYA LAKSHMI SWEETS", "JAYABAL STORE", "JAYADURGA STORE", "JAYALAKSHMI STORE(ARUM)", "JAYALAKSHMI SUPER  MARKET", "JAYAMANI STORE", "JAYARAM STORE", "JJ STORE(BAAI)", "JOSEPH STORE", "JUPITER SUPER MARKET", "KALA HOT CHIPS", "KALA STORE", "KALA SWEETS BAKEREY", "KAMATCHI STORE (ARUM)", "KAMATCHI STORE (KOYAM)", "KAMATCHI STORE (MMDA)", "KAMATCHI STORE(ANNA)", "KANAKADHAARA FRUITS AND VEGETABLES", "KANCHI FOODS", "KANDAN OIL STORE", "KANDAN VEG MART", "KANNAN STORE (ANNA)", "KANNIYAPPAN VEGETABLE", "KARIYA KALLIAMMAN FRUITS", "KARPAGA VINAYAKAR FRESH FRUIT", "KARPAGAM FOODS", "KARPAGAMBAAL FRESH FRUITS (ADYAR)", "KARPAGAMBAAL FRESH FRUITS (VALLUVARKOTTAM)", "KAVINS STORE", "KAVITHA HOSTEL", "KAVITHA STORE", "KAVITHA STORE (CHOO)", "KAVIYA & KAVITHA STORE", "KERALA HOT CHIPS & SNACKS", "KERALA HOT CHIPS & SNACKS (TAMBARAM)", "KERALA HOT CHIPS (SANTHI)", "KERALA HOT CHIPS AND SNACKS(COIM)", "KILPAUK GARDEN VEGEFURITS MARKET", "KING'S SUPER MARKET", "KIRTHIKA STORE", "KITCHA STORE (GB-84)", "KITCHAA STORE", "KOVAI PAZHAMUDHIR NILAYAM  (AMMAN)", "KPN FARM FRESH PRIVATE LIMITED(VANAGARAM)", "KRISHNA CAFE", "KSM CHETTIYAR CHEKKU OIL MILL", "KUMARAN STORE(CHOO)", "KUMARAN STORE(METHA)", "KURINJI TRADING POINT", "LAKSHMAN & LAKSHMI SWEETS", "LAKSHMI & CO", "LAKSHMI & CO (NEELANKARAI)", "LAKSHMI STORE (KOYM)", "LAKSHMI STORE (RV NAGAR)", "LAKSHMI STORE(VOC)", "LAKSHMI SWEETS", "LAKSHMI SWEETS (CHOO)", "LIMRAS ENTERPRISES (P) LTD", "LINGAM STORE", "LINGAM STORE (LOCAL)", "LINGAM STORE(MMDA)", "LUCKY STORE", "M. BAKKYAM STORE", "MAAHE SHOPPEE", "MADAN TRADERS", "MADINA TRADE POINT", "MADRAS -10", "MADRAS AGRO PRODUCTS", "MAHALAKSHI STORE", "MAHALAKSHMI ENTERPRISES", "MAHALAKSHMI OIL TRADERS", "MAHALAKSHMI PROVISION", "MAHALAKSHMI STORE", "MAHALAKSHMI STORE(CHOO)", "MAHARAJA STORE(CHOO)", "MANI TRADERS", "MANIGANDAN STORE", "MANO TRADERS", "MARGIN FREE SUPER MARKET", "MARINAKART", "MATHAJI STORE", "MEENAKSHI TRADERS", "MEGA MART", "MERCY HOME", "METRO GRAND HOTEL", "METRO SUPER MARKET", "MINAR BIRIYANI HOTEL", "MIRCHI HOME", "MIRORA HOTELS AND RESTAURANRTS (P).LTD", "MITIS SNACKS", "MOOGAMBIGAI STORE", "MURALI STORES", "MURUGAN STORE (18TH)", "MURUGAN STORE (AK 1)", "MURUGAN STORE (ARUM)", "MURUGAN STORE (G-87)", "MURUGAN STORE (K4)", "MURUGAN STORE (TP)", "MURUGAN STORE K4", "MURUGAN STORE(AYYA)", "MURUGAN STORE(JAI)", "MURUGAN STORE(KANDAN)", "MURUGAN STORE(KOYM)", "MURUGAN STORE(SHENOY)", "MURUGAN STORES SUPER MARKET (RAJAN)", "MURUGAN SWEETS", "MUTHU STORE", "MUTHUKRISHNA STORE", "MUTHUMALAI AMMAN STORE", "MUTTAI KADAI", "MY SHOP", "N.SELVAM (KMC)", "N.SUGANTHI (KMC)", "NALAS APPA KADAI", "NANDAN STORE", "NANDHINI TRADERS", "NATARAJA STORE", "NATIONAL STORE", "NATURE BOX TODAY", "NAVASAKTHI VINAYAGAR FRUITS", "NEW AYYANAR STORE", "NEW CHIDAMBARAM STORE", "NEW MAHE STORE(0)", "NEW NATIONAL STORE", "NEW PANDIAN STORE", "NEW PRABU STORE (GA32)", "NEW RAJA STORE", "NEW SAKTHI STORE", "NEW SHANTHI STORE", "NEW SHREE AYYANAR OIL STORE", "NEW THIRUMANI STORE", "NITHYA AMIRTHAM INDIAN FOOD PRIVATE LIMITED", "NITHYAM FOODS", "NSK HOT CHIPS", "NUTS 'N' SPICES", "OM MURUGA STORE", "OM STORE", "OM STORES(KOYM)", "P.R. AGENCIES", "P.S.P.FOODS AND GRAINS", "PADMA STORE", "PADMANABAN STORE", "PADMAVATHI SWEETS", "PALLAVI STORE", "PALRAJAN STORE", "PANDIAN STORE (ARUM)", "no use PANDIAN STORE (F-53)", "PANDIAN STORE (H-87)", "PANDIAN TRADERS", "PANDIAN TRADERS (F-53)", "PANDIAN TRADERS (GC-25)", "PANDIYAN TRADERS (D-86)", "PANJALI STORE", "PATHRAKALI STORE", "PATHUBANATHAN STORE", "PATTU RAJA STORE", "PATTU STORE (GA-36)", "PATTU STORE (GA-36)", "PAZHAMUDHIR NILAYAM", "PAZHAMUDIR NILAYAM FARM FRESH (GOPALAPURAM)", "PAZHAMUDIR NILAYAM FARM FRESH(MANDAVELI)", "PERIYANDAVAR OIL STORE", "PERIYANDAVAR STORE", "PERIYAPERATTI STORE", "PIECE OF CAKE", "PKR STORE", "PONMANI SWEETS", "PONNAIAMMAN STORE", "PONNUSAMY HOTEL", "PRABHA STORE MMDA", "PRABU STORE (A-121)", "PRABU STORE (GA-32)", "PRAGATI FOODS", "PRAKASH STORE", "QUALITY HOME", "QUIET SUPER MARKET", "R.K. NITISH STORE", "R.V TRADERS", "RABBANA STORE", "RAGAVENDRA STORE", "RAJA OIL STORE", "RAJA STORE (ARUM)", "RAJA STORE(ANNA)", "RAJA STORE(KOYM)", "RAJAM STORE", "RAJAMMAL ENTERPRISES", "RAJAPPAN STORE", "RATHI STORE", "RBN STORES", "REDDY ANDHRA MESS", "RENGA AGENCY", "RENGA STORE", "REVATHI STORE", "REVATHI SUPER MARKET", "ROBERT STORE(TP)", "RR AGENCY", "RT GURUSAMY STORE", "RV STORE", "S.S. GREEN BASKET", "S.S.S.STORE", "SA SWAMY STORE", "SADHANA TRADERS", "SAI PROVISION STORE", "SAI RAM CHIPS", "SAI RAM HOT CHIPS", "SAI SUPER MARKET", "SAIS PROVISION STORE", "SAKTHI STORE (ANNA )", "SAKTHI STORE (KAT)", "SAKTHI SUPER MARKET", "SAKTHI VINAYAGAM STORE", "SAMBATH STORE", "SANABEL BRIYANI POINT", "SANDHYA BHAVAN HOTEL", "SANKAR STORE", "SANKARAN AGENCIES", "SANTHOSH STORE(ARUM)", "SANTHOSH SUPER MARKET(ANNA)", "SANTHOSH SUPER STORE", "SANTHOSH SUPER STORE(18TH)", "SARAN SWEETS", "SARANYA STORE(CHOO)", "SARASWATHI RICE STORE", "SARAVANA AGENCIES", "SARAVANA STORE (18 TH)", "SARAVANA STORE (KOYM)", "SARAVANA STORE (PARK RD)", "SARAVANA STORE (Y 20 )", "SARAVANA STORE(18TH)", "SARAVANA STORE(KANDAN)", "SARAVANA STORE(TP)", "SARAVANAS SUPER MARKET", "SASIKUMAR STORE", "SATHANA TRADERS", "SCARLATE SUPER MARKET", "SEDHU STORE", "SEKAR STORE", "SEKAR STORES (AMMAN)", "SELLAPPA HIGH QUALITY VEG", "SELLIAMMAN STORE", "SELVA GANESH STORE", "SELVAI STORE", "SELVAKUMAR PROVISION STORE", "SELVAKUMAR SUPER MARKET", "SELVAM OIL STORE (NEW)", "SELVAM STORE", "SELVAM STORE (7TH)", "SELVAM STORE (MADURAVOYAL)", "SELVAM STORE (NEDUVAKARAI)", "SELVAM STORE (NSK)", "SELVAM STORE(ANNA)", "SELVAM STORE(ARUM)", "SELVAM STORE(LAKSHMAN)", "SELVAM STORE(THANGAM)", "SELVAM STORE(THIRU MULAR)", "SELVAM SWEETS & BAKERY(SHENOY)", "SELVAM SWEETS BAKERY AND SNACKS", "SELVI SUPER MARKET", "SELVI SUPER MARKET (SHENOY)", "SENTHIL ANDHAVAR FRUITS", "SENTHIL ANDHAVAR STORE", "SENTHIL MURUGAN STORE", "SENTHIL TRADERS", "SENTHIL VEL STORE", "SH STORE", "SHAHI FOODS", "SHANDHI OIL MILLS", "SHANMUGASUNDARAM STORE", "SHANTHI OIL", "SHOPPING SINGAPPORE (NEW)", "SHREE AKHILANDESWARI AMMAN FRUITS", "SHREE RAJAVELL TRADERS", "SHRI SABARI AGENCY", "SHRI SARASWATHI AMMA FARM FRUITS", "SHRIYAA CHATS AND FOODS", "SIDDHA LADIES HOSTEL", "SIDDHESWARA & CO", "SIDDHI KRIPA", "SIVA RAMA TRADERS", "SIVA RATHINAM TRADERS", "SIVA STORE (10TH)", "SIVA STORE (NSK)", "SIVA STORE (THANGA COLONY)", "SIVAN STORE", "SIVARATHINAM STORE(GA-34)", "SIVAYOGAM & CO", "SK STORE", "SKS MINI SUPER MARKET", "SM DURAI RAJ STORE(A-86)", "SM DURAI RAJ STORE(E87)", "SM DURAIRAJ STORE(A-91)", "SM TRADERS", "SNEHA STORE", "SNV VENTURES LLP", "SORNA LAKSHMI", "SOWBAGHYA", "SOWTHA BAKERY", "SOWTHA STORE", "SR DEPARTMENT STORE", "SREE BHARANI FAST FOOD", "SREE JAYALAKSHMI STORE", "SREE KALA HOT CHIPS", "SREE KARAN ENTERPRISES", "SREE KRISHNA PROVISION STORE", "SREE LATHA STORE", "SREE MAHALAKSHMI MARKETING", "SREE NIVAS UDUPI HOTEL", "SREE RADHA & CO", "SREE RAJARAJESWARI RICE TRADING COMPANY", "SREE RAM TRADERS", "SRI AKSHAYAAS HOTEL", "SRI AYYANAR STORE (D27)", "SRI AYYANAR TRADERS (37)", "SRI AYYANAR TRADERS(GA307)", "SRI BALAJI STORE", "SRI BALAJI TRADING COMPANY", "SRI BALASUBRAMANIYAR TRADERS", "SRI BANNARI AMMAN FRESH FRUIT", "SRI CATERING", "SRI GANESH STORE (JAI)", "SRI HARI HARAN ENTERPRISES", "SRI JAI DURGA STORE", "SRI KANDAN OIL STORE", "SRI KRISHNA PROVISION", "SRI LAKSHMI STORE", "SRI MAHALAKSHMI STORE", "SRI MURUGAN ENTERPRISES", "SRI MURUGAN STORE (CHOO)", "SRI MURUGAN STORE(SHENOY)", "SRI PAZHANI STORE", "SRI RAGAVENDRA STORE", "SRI RAGAVENDRA TRADERS", "SRI RAMAJAYAM STORE", "SRI RANGA TRADERS", "SRI SAI SABARI STORE", "SRI SAKTHI STORE", "SRI SARAPAR MURUGAN STORE", "SRI SATHYA SUPER MARKET", "SRI SRINIVASA STORE", "SRI VARI HOT CHIPS", "SRI VIJAYLAKSHMI STORE", "SRINITHI HOTEL", "SRINIVASA STORE", "SRINIVASAN", "SRIVARI SWEETS", "SSK TRADERS", "SUBA STORE(TP)", "SUHAIL ENTERPRISES (P). LTD.", "SUMATHI STORE(TP)", "SUN OIL TRADERS", "SUN SUPER STORE", "SUNDARAM STORE(AMINI)", "SUNDARAM STORE(ARUM)", "SUNDARI STORES", "SUPRABHA HOTEL (0)", "SYED STORE", "TAMILNADU (NATTUMARUNTU KADAI)", "TAMILNADU STORE", "TAYLORS  ROAD (VEG & FRUITS)", "THAI STORE(CHOO)", "THANGAM PROVISION STORE", "THANGAM STORE", "THANGAM STORE(CMBT)", "THANGAM STORE(KOYM)", "THANGAMANI STORE", "THANGAPANDIYAN STORE", "THANGAPPAN STORE", "THANJAI PROVISION STORE", "THE CREAMS", "THE FOODLE STORE", "THE GRAND SWEETS AND SNACKS", "THILAGAM STORE", "THIRUKUMARAN STORE", "THIRUMALA STORE", "THIRUMALA SUPER MARKET", "THIRUPATHI STORE", "THIRUVENI HOTEL", "THIRUVENI STORE", "UDUPI GANESH BHAVAN", "UDIPI SANKAR BHAVAN", "V S STORE", "V.R. CATERING SERVICE", "V.S. TRADERRS", "VAGAP STORE", "VAISHNAVI SWEETS", "VANAJA STORE", "VARADHARAJAN STORE(C-52)", "VARADHARAJAN TRADERS(GC-9)", "VARATHARAJAN STORE(E-86)", "VARATHARAJAN STORE(E-87)", "VASANT CAFE", "VASANTHA BHAVAN HOTEL", "VASEER STORE", "VEERA VEG MART", "VEL STORE (ANNA)", "VELUS PAZHAMUDIR SOLAI & SUPER MARKET", "VENKATESWARA CHIPS(ANNA)", "VENKATESWARA SWEETS", "VENNILA STORE", "VETTA SUPER STORE", "VETTRII & CO", "VIGNESH CATERING SERVICE", "VIGRAMA HOTEL", "VIJAYA STORE", "VIJAYALAKSHMI STORE", "VIJAYALAKSHMI SWEETS", "VIJAYALAKSHMI TRADERS", "VIKPRI HOME NEEDS", "VINAYAGA GENERAL STORES", "VINAYAGA HOT CHIPS", "VINAYAGA HOTEL", "VINAYAGA HOTEL BHAVAN", "VISHNU PRIYA STORE", "VOICE NATURAL AND FOOD PRODUCTS", "VOPEC PHARAMACEUTICALS", "VP SHOPPING", "VPD ENTERPRISES", "VRP TRADERS", "WINWAY FOOD PRODUTS", "XAVIER STORE", "YASHI PROVISIONAL TRADERS", "YESESI SUPER MARKET", "YOVAN STORE", "ZAKIR STORES", "AK SUPER MARKET", "AV STORES", "AV STORE(2)", "A V STORES 2", "AP SELVAM STORE", "ARCHANA STORE", "ANNAI STORE(BELL)", "AROKYA BHAVAN HOTEL", "SUDALAI", "AKSHAYA TRADERS", "MURUGAN STORE (AMINJI) RAMAN", "AL-KHADER PROVSION(NEW)", "RAJA DEPARTMENTAL STORE", "KRISHNA MINI MART", "SRITHAR AGENCIES", "SIDDHESWARA & CO", "NEW SELVAKUMAR STORE(NEW)", "ARFAA HOTEL", "idhyam order", "GRB ORDER", "COFFEE HOUSE AT MOUNT ROAD", "ANANDA STORE (AYYAVO COLONY)", "MURUGAN SUPER MARKET (CHOOLAI)", "KGN HOSTEL", "PANI POORI", "ANNAM,HITAM ASSOCIATES PVT LTD", "TAYLORS  ROAD (Oil & Rice)", "amen moiden store", "RAJAN STORES", "Orange super market", "REKHAA ORGANICS (FRESH2DAY)", "FRESH N FAIR", "MANGALAM SUPER MARKET",]
area_dict = {"Choolaimedu":Choolaimedu,"Anna Nagar":anna_nagar,"Full":l,"MMDA":mmda, "Arumbakkam":arumbakkam}
area_count = {"Choolaimedu":0,"Anna Nagar":0,"Full":0,"MMDA":0,"Arumbakkam":0}
completed = []

def get_list(e):
    total.config(text="Total collection of %s: %d" % (drop_down.get(), area_count[drop_down.get()]))
    l = area_dict[drop_down.get()]
    list.config(height=10)
    out = []
    leng = len(str(data.get()))
    if data.get() != "":
        for i in l:
            if(data.get().upper().startswith(i[:leng])):
                out.append(i)
        list.delete(0,END)
    else:
        out = l[:]
        list.delete(0, END)
    for i in out:
        list.insert(END,i)
        if i in completed:
            list.itemconfig(END,bg ="green",selectbackground="green")
    global store_name
    store_name=list.select_set(0)


def enter(e):
    list.config(height=10)
    list.focus_set()


def change_value(ev):
    output_frame.grid_slaves(row=1,column=3)[0].focus_set()


def show_accounts(ev):
    out = []
    list.config(height=7)
    global store_name
    for i in list.curselection():
        store_name = list.get(i)
        wb = load_workbook(store_name+".xlsx")
        pointer = wb.active
        for i in output_frame.winfo_children():
            i.destroy()
        Label(output_frame, text="Store Name:", background='light grey',font=("Calibri", 15)).grid(row=0,column=0)
        Label(output_frame, text=store_name, font=("Calibri", 15)).grid(row=1,column=0)
        Label(output_frame, text="Date:", background='light grey', font=("Calibri", 15)).grid(row=0, column=1, padx=100)
        Label(output_frame, text="Amount:", background='light grey',font=("Calibri", 15)).grid(row=0,column=2)
        Label(output_frame, text="Amt Received:", background='light grey', font=("Calibri", 15)).grid(row=0, column=3)
        Label(output_frame, text="Balance:", background='light grey', font=("Calibri", 15)).grid(row=0, column=4)
        date = pointer["A"]
        amt = pointer["B"]
        for row,i in enumerate(amt):
            if type(i.value) == int:
                Label(output_frame, text = date[row].value,font=("Calibri", 15)).grid(row=row+1, column=1)
                Label(output_frame, text = i.value, font=("Calibri", 15)).grid(row=row+1,column=2)
                e = Entry(output_frame,relief=GROOVE,font=("Calibri", 15))
                e.grid(row=row+1, column=3,padx=80,pady=2)
                e.bind("<KeyRelease>", calculate)
                e.bind("<Down>", movement_down)
                e.bind("<Up>", movement_up)
                e.bind("<Return>", movement_down)
                e.bind("<Escape>",enter)
                #e.insert(END,i.value)
                Label(output_frame,font=("Calibri", 15)).grid(row=row + 1, column=4)
                output_frame.pack(padx=2,pady=2)
        e.bind("<Return>", jump_to_save)

def calculate(ev):
    new = output_frame.focus_get().grid_info()['row']
    val = output_frame.focus_get().get()
    if val.isnumeric():
        bal = output_frame.grid_slaves(row=new, column=2)[0].cget("text") - int(val)
        output_frame.grid_slaves(row=new,column=4)[0].configure(text=bal)
    else:
        output_frame.grid_slaves(row=new, column=4)[0].configure(text=output_frame.grid_slaves(row=new, column=2)[0].cget("text"))


def movement_down(e):
    total_row =output_frame.grid_size()[1]
    info = output_frame.focus_get().grid_info()
    current_row = info["row"]+1
    if total_row-1 >= current_row:
        output_frame.grid_slaves(current_row,column=3)[0].focus_set()


def movement_up(e):
    info = output_frame.focus_get().grid_info()
    current_row = info["row"]-1
    if 1 <= current_row:
        output_frame.grid_slaves(current_row, column=3)[0].focus_set()


def jump_to_save(e):
    buttons_of_operation.grid_slaves(row=0, column=2)[0].focus_set()


def pop_ask(e):
    def yes_even(ev):
        print("saved")
        save_to_main()
        pop.destroy()

    def no_event(ev):
        print("not saved")
        text.focus_set()
        pop.destroy()

    def highlight_yes(ev):
        yes.focus_set()

    def highlight_no(ev):
        no.focus_set()

    global pop
    pop = Toplevel(root)
    pop.geometry("250x100")
    pop.title("Save?")
    Label(pop, text="Do you really want to save?").grid(row=0,column=1)
    yes = Button(pop,text="Yes")
    yes.bind("<Return>", yes_even)
    yes.bind("<Right>", highlight_no)
    yes.focus_set()
    yes.grid(row=1,column=1,pady=10)
    no = Button(pop,text="No")
    no.grid(row=1,column=2,pady=10)
    no.bind("<Return>", no_event)
    no.bind("<Left>", highlight_yes)
    pop.grid()

def save_to_main():
    print(store_name)
    wb = load_workbook(str(store_name) + ".xlsx")
    pointer = wb.active
    total_row = output_frame.grid_size()[1]
    pointer.delete_cols(2)
    insert_pointer = 1
    for i in range(1,total_row):
        value = output_frame.grid_slaves(row=i,column=4)[0].cget("text")
        balance = output_frame.grid_slaves(row=i,column=3)[0].get()
        print(value, i)
        if balance.isnumeric():
            if drop_down.get() != "Full":
                area_count["Full"] += int(balance)
            area_count[drop_down.get()] += int(balance)
            total.config(text="Total collection of %s: %d" % (drop_down.get(),area_count[drop_down.get()]))
        if value != 0:
            pointer['B%d' % insert_pointer].value = int(value)
            insert_pointer += 1
        else:
            pointer.delete_rows(i)
    wb.save((store_name) + ".xlsx")
    completed.append(store_name)
    text.delete(0,END)
    escape_operation()


def escape_operation(ev=0):
    text.focus_set()
def view_store():
    l = area_dict[drop_down.get()]
    book = openpyxl.Workbook()
    p = book.active
    for i in l:
        wb = load_workbook("%s.xlsx" % i)
        pointer = wb.active
        date = pointer["A"]
        amt = pointer["B"]
        for row, j in enumerate(amt):
            if type(j.value) == int:
                #print(i+":", j.value, date[row].value,"\n")
                p.append((i, date[row].value, j.value))
    file_name = "_"+drop_down.get()+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    print(file_name)
    book.save(str(file_name)+".xlsx")
    book.save(os.path.dirname(__file__)+"//Records//"+str(file_name) + ".xlsx")


Label(root, text="WELCOME TO COLLECTION SOFTWARE").pack(side=TOP,pady=5)

input_frame = Frame(root)
Label(input_frame,text= "Store Name : ").grid(row=0,column=0)
data = StringVar()
text = Entry(input_frame,width=90,bd=15,textvariable=data)
text.bind("<KeyRelease>",get_list)
text.bind("<Return>", enter)
text.bind("<Down>", enter)
text.grid(row=0,column=1)
text.focus_set()
drop = StringVar(root)
drop.set(area_list[0])
drop_down = Combobox(input_frame,value = area_list,font=("Calibri", 15))
drop_down.current(0)
drop_down.grid(row=0, column=2,padx=10)
drop_down.bind('<<ComboboxSelected>>', get_list)
drop_down.bind('<Escape>', escape_operation)
input_frame.pack(pady=10)


list = Listbox(root, height=10,font=("Calibri", 15))
list.pack(fill='x',pady=5)
list.bind("<KeyRelease>",show_accounts)
list.bind("<Double-Button-1>", show_accounts)
list.bind("<Return>", change_value)
list.bind("<Escape>", escape_operation )
store_name = ""

scroll = Scrollbar(root)

output_frame = Frame(root,highlightbackground="black", highlightthickness=2 )


buttons_of_operation = Frame(root)
total = Label(buttons_of_operation,text = "Total collection of %s: %d" % (drop_down.get(),area_count[drop_down.get()]),font=("Calibri", 15))
total.grid(row=0,column=0,padx=15)
view = Button(buttons_of_operation, text="View", command=view_store).grid(row=0,column=1,padx=15)
save_but = Button(buttons_of_operation, text="Save Changes")
save_but.bind("<Return>", pop_ask)
save_but.bind("<Button-1>", pop_ask)
save_but.grid(row=0,column=2,padx=25)
buttons_of_operation.pack(side=BOTTOM,pady=(0,50))


lab = Label(root,text="")
lab.pack()

root.mainloop()
