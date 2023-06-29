from tkinter import *

from openpyxl import load_workbook

root = Tk()
root.geometry("1080x720")

l = ['sasi','karan','jeya','sivakumar']


def get_list(e):
    list.config(height=10)
    out = []
    leng = len(str(data.get()))
    if data.get() != "":
        for i in l:
            if(data.get().startswith(i[:leng])):
                out.append(i)
        list.delete(0,END)
    else:
        out = l[:]
        list.delete(0, END)
    for i in out:
        list.insert(END,i)
    global store_name
    store_name=list.select_set(0)


def enter(e):
    list.config(height=10)
    list.focus_set()

def change_value(ev):
    output_frame.grid_slaves(row=1,column=1)[0].focus_set()

def show_accounts(ev):
    out = []
    list.config(height=7)
    global store_name
    for i in list.curselection():
        store_name = list.get(i)
        wb = load_workbook(store_name+".xlsx")
        pointer = wb.active
        for i in output_frame.winfo_children():
            i.destroy()
        Label(output_frame, text="Store Name:",background='light grey',font=("Calibri", 15)).grid(row=0,column=0)
        Label(output_frame, text=store_name,font=("Calibri", 15)).grid(row=1,column=0)
        Label(output_frame, text="Amount:",background='light grey',font=("Calibri", 15)).grid(row=0,column=1)

        for row,i in enumerate(pointer["B"]):
            e = Entry(output_frame,relief=GROOVE,font=("Calibri", 10))
            e.grid(row=row+1, column=1,padx=200,pady=2)
            e.bind("<Down>", movement_down)
            e.bind("<Up>", movement_up)
            e.bind("<Return>", movement_down)
            e.bind("<Escape>",enter)
            e.insert(END,i.value)
            output_frame.pack(padx=2,pady=2)
            if row+1 == len(pointer["B"]):
                e.bind("<Return>", jump_to_save)


def movement_down(e):
    total_row =output_frame.grid_size()[1]
    info = output_frame.focus_get().grid_info()
    current_row = info["row"]+1
    if total_row-1 >= current_row:
        output_frame.grid_slaves(current_row)[0].focus_set()

def movement_up(e):
    info = output_frame.focus_get().grid_info()
    current_row = info["row"]-1
    if 1 <= current_row:
        output_frame.grid_slaves(current_row)[0].focus_set()

def jump_to_save(e):
    buttons_of_operation.grid_slaves(row=0,column=1)[0].focus_set()

def pop_ask(e):

    def yes_even(e):
        print("saved")
        save_to_main()
        pop.destroy()

    def no_event(e):
        print("not saved")
        text.focus_set()
        pop.destroy()

    def highlight_yes(e):
        yes.focus_set()

    def highlight_no(e):
        no.focus_set()

    global pop
    pop = Toplevel(root)
    print("poped")
    pop.geometry("250x100")
    pop.title("Save?")
    Label(pop, text="Do you really want to save?").grid(row=0,column=1)
    yes = Button(pop,text="Yes")
    yes.bind("<Return>", yes_even)
    yes.bind("<Right>", highlight_no)
    yes.focus_set()
    yes.grid(row=1,column=1,pady=10)
    no = Button(pop,text="No")
    no.grid(row=1,column=2,pady=10)
    no.bind("<Return>", no_event)
    no.bind("<Left>", highlight_yes)
    pop.grid()

def save_to_main():
    print(store_name)
    wb = load_workbook(str(store_name) + ".xlsx")
    pointer = wb.active
    total_row = output_frame.grid_size()[1]
    new = []
    for i in range(1,total_row):
        value = output_frame.grid_slaves(i)[0].get()
        pointer['C%d' % i].value = value
    wb.save((store_name) + ".xlsx")
    # l.remove(store_name)
    text.delete(0,END)
    escape_operation()

def escape_operation(ev=0):
    text.focus_set()
def view_store():
    canva = Canvas(root,height=200,width=200)


Label(root, text="WELCOME TO COLLECTION SOFTWARE").pack(side=TOP,pady=5)

input_frame = Frame(root)
Label(input_frame,text= "Store Name : ").grid(row=0,column=0)
data = StringVar()
text = Entry(input_frame,width=90,bd=15,textvariable=data)
text.bind("<KeyRelease>",get_list)
text.bind("<Return>", enter)
text.bind("<Down>", enter)
text.grid(row=0,column=1)
input_frame.pack(pady=10)


list = Listbox(root, height=10,font=("Calibri", 15))
list.pack(fill='x',pady=5)
list.bind("<KeyRelease>",show_accounts)
list.bind("<Double-Button-1>", show_accounts)
list.bind("<Return>", change_value)
list.bind("<Escape>", escape_operation )
store_name = ""

scroll = Scrollbar(root)

output_frame = Frame(root,highlightbackground="black", highlightthickness=2 )


buttons_of_operation = Frame(root)
view = Button(buttons_of_operation, text="View", command=view_store).grid(row=0,column=0,padx=5)
save_but = Button(buttons_of_operation, text="Save Changes")
save_but.bind("<Return>", pop_ask)
save_but.bind("<Button-1>", pop_ask)
save_but.grid(row=0,column=1,padx=5)
buttons_of_operation.pack(side=BOTTOM)


lab = Label(root,text="")
lab.pack()

root.mainloop()