import openpyxl

workspace = openpyxl.load_workbook("sasi.xlsx")
sheet = workspace.active
x = sheet.cell(1, 1).value
print(x)
cell=sheet["B1"]
cell.value = "karan"
#sheet.append(["karan"])
print(sheet.max_row)
workspace.save("sasi.xlsx")


